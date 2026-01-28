from datetime import datetime
from typing import Dict, Any, List

import os
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib.enums import TA_CENTER
except ImportError:
    print("Missing dependencies: openpyxl or reportlab")

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        width, height = self._pagesize
        self.drawCentredString(width / 2.0, 0.7 * cm,
                               "Page %d / %d" % (self._pageNumber, page_count))

def format_currency(value):
    return f"{value:,.2f} DA".replace(",", " ").replace(".", ",")

def format_currency_report(value):
    if value is None: return "0,00"
    try:
        val = float(value)
        return f"{val:,.2f}".replace(",", " ").replace(".", ",")
    except (ValueError, TypeError):
        return str(value)

def generate_stock_valuation_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etat Stock Valorise"
    
    # Styles
    bold_font = Font(bold=True, name='Arial', size=10)
    title_font = Font(bold=True, name='Arial', size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Title
    ws['B2'] = "ETAT DES MOUVEMENTS DES STOCKS (VALORISES)"
    ws['B2'].font = title_font
    ws['B2'].alignment = center_align
    ws.merge_cells('B2:K2')
    
    # Meta Data
    prod_name = data['product']['nom']
    start_date = data['period']['start']
    end_date = data['period']['end']
    unit = data['product']['unite']
    
    ws['A3'] = f"PRODUIT : {prod_name}"
    ws['A3'].font = bold_font
    
    ws['D3'] = f"DU : {start_date}"
    ws['D3'].font = bold_font
    
    ws['F3'] = f"AU : {end_date}"
    ws['F3'].font = bold_font
    
    ws['I3'] = f"Date d'établissement : {datetime.now().strftime('%d/%m/%Y')}"
    
    ws['A4'] = f"UNITE DE MESURE : {unit}"
    ws['A4'].font = bold_font
    
    # Headers
    # Row 6
    headers_row6 = [
        ("A6", "JOURNEE"),
        ("B6", "STOCK INITIAL"), 
        ("D6", "P.UNITAIRE"),
        ("E6", "RECEPTIONS"),
        ("G6", "VENTES"),
        ("I6", "STOCK FINAL")
    ]
    
    for cell_ref, text in headers_row6:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        
    # Merges
    ws.merge_cells('A6:A7') # Journee merges vertically
    ws.merge_cells('B6:C6') # Stock Initial merges horizontally
    # ws.merge_cells('D6:D7') # REMOVED: P.Unit and C.Achat are separate cells vertically
    ws.merge_cells('E6:F6')
    ws.merge_cells('G6:H6')
    ws.merge_cells('I6:J6')
    
    # Correcting merges based on logic
    # REMOVED DUPLICATES
    
    # Row 7 Sub-headers
    headers_row7 = [
        ("B7", "QUANTITES"), ("C7", "VALEURS"),
        ("D7", "C.ACHAT"),
        ("E7", "QUANTITES"), ("F7", "VALEURS"),
        ("G7", "QUANTITES"), ("H7", "VALEURS"),
        ("I7", "QUANTITES"), ("J7", "VALEURS"),
    ]
    
    for cell_ref, text in headers_row7:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border

    # Data
    row_idx = 8
    for row_data in data['data']:
        # Format date
        date_obj = datetime.strptime(row_data['date'], '%Y-%m-%d')
        date_fmt = date_obj.strftime('%d/%m/%Y')
        
        # Columns:
        # A: Date
        # B: Stock Init Qty
        # C: Stock Init Val
        # D: C.Achat
        # E: Recep Qty
        # F: Recep Val
        # G: Vente Qty
        # H: Vente Val
        # I: Final Qty
        # J: Final Val
        
        ws.cell(row=row_idx, column=1, value=date_fmt).border = border
        
        ws.cell(row=row_idx, column=2, value=format_currency_report(row_data['stock_initial_qty'])).border = border
        ws.cell(row=row_idx, column=3, value=format_currency_report(row_data['stock_initial_val'])).border = border
        
        ws.cell(row=row_idx, column=4, value=format_currency_report(row_data['cout_achat'])).border = border
        
        ws.cell(row=row_idx, column=5, value=format_currency_report(row_data['reception_qty'])).border = border
        ws.cell(row=row_idx, column=6, value=format_currency_report(row_data['reception_val'])).border = border
        
        ws.cell(row=row_idx, column=7, value=format_currency_report(row_data['vente_qty'])).border = border
        ws.cell(row=row_idx, column=8, value=format_currency_report(row_data['vente_val'])).border = border
        
        ws.cell(row=row_idx, column=9, value=format_currency_report(row_data['stock_final_qty'])).border = border
        ws.cell(row=row_idx, column=10, value=format_currency_report(row_data['stock_final_val'])).border = border
        
        row_idx += 1
        
    # Footer
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="LE CHEF SERVICE COMMERCIAL").font = bold_font
    ws.cell(row=row_idx, column=5, value="LE CHEF SERVICE COMPTABILITE").font = bold_font
    ws.cell(row=row_idx, column=8, value="LE CHEF DU DEPOT").font = bold_font
    
    # Column Widths
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 12

    wb.save(output_path)
    return output_path

def generate_stock_valuation_pdf(data, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    title = Paragraph("ETAT DES MOUVEMENTS DES STOCKS (VALORISES)", title_style)
    elements.append(title)
    elements.append(Spacer(1, 1*cm))
    
    # Meta Data
    prod_name = data['product']['nom']
    start_date = data['period']['start']
    end_date = data['period']['end']
    unit = data['product']['unite']
    
    meta_data = [
        [f"PRODUIT : {prod_name}", f"DU : {start_date}   AU : {end_date}", f"Editée le : {datetime.now().strftime('%d/%m/%Y')}"],
        [f"UNITE DE MESURE : {unit}", "", ""]
    ]
    t_meta = Table(meta_data, colWidths=[8*cm, 10*cm, 8*cm])
    t_meta.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 0.5*cm))
    
    # Table Data
    # Headers
    # We need to simulate merged cells by managing the grid manually or using span
    # ReportLab Table allows span
    
    # Header Row 1
    h1 = [
        "JOURNEE", 
        "STOCK INITIAL", "", 
        "P.UNITAIRE", 
        "RECEPTIONS", "", 
        "VENTES", "", 
        "STOCK FINAL", ""
    ]
    
    # Header Row 2
    h2 = [
        "", # Under Journee
        "QUANTITES", "VALEURS", 
        "C.ACHAT", # Under P.Unit
        "QUANTITES", "VALEURS",
        "QUANTITES", "VALEURS",
        "QUANTITES", "VALEURS"
    ]
    
    table_data = [h1, h2]
    
    t_recept_qty = 0.0
    t_recept_val = 0.0
    t_vente_qty = 0.0
    t_vente_val = 0.0

    for row in data['data']:
        date_obj = datetime.strptime(row['date'], '%Y-%m-%d')
        date_fmt = date_obj.strftime('%d/%m/%Y')
        
        # Accumulate Totals
        t_recept_qty += float(row['reception_qty'])
        t_recept_val += float(row['reception_val'])
        t_vente_qty += float(row['vente_qty'])
        t_vente_val += float(row['vente_val'])

        table_data.append([
            date_fmt,
            format_currency_report(row['stock_initial_qty']),
            format_currency_report(row['stock_initial_val']),
            format_currency_report(row['cout_achat']),
            format_currency_report(row['reception_qty']),
            format_currency_report(row['reception_val']),
            format_currency_report(row['vente_qty']),
            format_currency_report(row['vente_val']),
            format_currency_report(row['stock_final_qty']),
            format_currency_report(row['stock_final_val']),
        ])
        
    # TOTAL ROW
    total_row = [
        "TOTAL",
        "", "", # Stock Init (No Total)
        "",     # P.Unit (No Total)
        format_currency_report(t_recept_qty),
        format_currency_report(t_recept_val),
        format_currency_report(t_vente_qty),
        format_currency_report(t_vente_val),
        "", "", # Stock Final (No Total)
    ]
    table_data.append(total_row)

    t = Table(table_data, colWidths=[2.5*cm] + [2.5*cm]*9)
    
    # Base Styles
    base_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'), # GLOBAL BOLD
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        
        # Merges
        ('SPAN', (0,0), (0,1)), # Journee
        ('SPAN', (1,0), (2,0)), # Stock Initial
        # ('SPAN', (3,0), (3,1)), # REMOVED
        ('SPAN', (4,0), (5,0)), # Receptions
        ('SPAN', (6,0), (7,0)), # Ventes
        ('SPAN', (8,0), (9,0)), # Stock Final
        
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        # Style Total Row (Last Row)
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
    ]
    
    # Conditional Styles (Skip 2 header rows, Col 0 is Date - handled by parser)
    cond_styles = get_conditional_styles(table_data[2:], start_row=2, start_col=0)
    
    t.setStyle(TableStyle(base_styles + cond_styles))
    
    elements.append(t)
    elements.append(Spacer(1, 2*cm))
    
    # Footer
    footer_data = [
        ["LE CHEF SERVICE COMMERCIAL", "", "LE CHEF SERVICE COMPTABILITE", "", "LE CHEF DU DEPOT"]
    ]
    t_foot = Table(footer_data, colWidths=[6*cm, 2*cm, 6*cm, 2*cm, 6*cm])
    t_foot.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t_foot)
    
    doc.build(elements)
    return output_path

def generate_global_consumption_excel(date_str, output_path=None):
    from logic import get_logic
    logic = get_logic()
    data = logic.get_global_consumption_data(date_str)
    
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Etat_Conso_Global_{date_str}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etat Consommation"
    
    # Styles
    bold_font = Font(bold=True, name='Arial', size=10)
    title_font = Font(bold=True, name='Arial', size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Title
    ws['B2'] = f"ETAT DE CONSOMMATION GLOBAL - JOURNEE DU {datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    ws['B2'].font = title_font
    ws['B2'].alignment = center_align
    ws.merge_cells('B2:J2')
    
    # Headers
    headers = [
        ("A4", "Désignation"), ("B4", "U"),
        ("C4", "JOURNEE"), ("E4", "CUMUL MOIS"), ("G4", "CUMUL ANNEE")
    ]
    
    # Sub-headers
    sub_headers = [
        ("C5", "Qté"), ("D5", "Valeur"),
        ("E5", "Qté"), ("F5", "Valeur"),
        ("G5", "Qté"), ("H5", "Valeur")
    ]
    
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border

    for cell_ref, text in sub_headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        
    # Merges
    ws.merge_cells('A4:A5')
    ws.merge_cells('B4:B5')
    ws.merge_cells('C4:D4')
    ws.merge_cells('E4:F4')
    ws.merge_cells('G4:H4')
    
    row_idx = 6
    for row in data['data']:
        # A: Name, B: Unit, C: Day Q, D: Day V, E: Month Q, F: Month V, G: Year Q, H: Year V
        
        ws.cell(row=row_idx, column=1, value=row['product_name']).border = border
        ws.cell(row=row_idx, column=2, value=row['unit']).border = border
        ws.cell(row=row_idx, column=2).alignment = center_align
        
        ws.cell(row=row_idx, column=3, value=format_currency_report(row['daily_qty'])).border = border
        ws.cell(row=row_idx, column=4, value=format_currency_report(row['daily_val'])).border = border
        
        ws.cell(row=row_idx, column=5, value=format_currency_report(row['monthly_qty'])).border = border
        ws.cell(row=row_idx, column=6, value=format_currency_report(row['monthly_val'])).border = border
        
        ws.cell(row=row_idx, column=7, value=format_currency_report(row['yearly_qty'])).border = border
        ws.cell(row=row_idx, column=8, value=format_currency_report(row['yearly_val'])).border = border
        
        row_idx += 1
        
    # Signatures
    row_idx += 3
    ws.cell(row=row_idx, column=2, value="Section Facturation").font = bold_font
    ws.cell(row=row_idx, column=7, value="Chef Service Commercial").font = bold_font

    # Column Widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

    wb.save(output_path)
    return output_path

def generate_global_consumption_pdf(date_str, output_path=None):
    from logic import get_logic
    logic = get_logic()
    data = logic.get_global_consumption_data(date_str)
    
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Etat_Conso_Global_{date_str}_{timestamp}.pdf"

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Logo Check
    from utils import check_logo_exists
    if check_logo_exists():
        try:
            from reportlab.platypus import Image as RLImage
            im = RLImage("logo_gica.png", width=4*cm, height=2*cm)
            im.hAlign = 'LEFT'
            elements.append(im)
        except: pass
        
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    date_fmt = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    title = Paragraph(f"ETAT DE CONSOMMATION GLOBAL - A FIN {date_fmt}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 1*cm))
    
    # Table Header
    h1 = ["Désignation", "U", "JOURNEE", "", "CUMUL MOIS", "", "CUMUL ANNEE", ""]
    h2 = ["", "", "Qté", "Valeur", "Qté", "Valeur", "Qté", "Valeur"]
    
    table_data = [h1, h2]
    
    for row in data['data']:
        table_data.append([
            row['product_name'],
            row['unit'],
            format_currency_report(row['daily_qty']),
            format_currency_report(row['daily_val']),
            format_currency_report(row['monthly_qty']),
            format_currency_report(row['monthly_val']),
            format_currency_report(row['yearly_qty']),
            format_currency_report(row['yearly_val'])
        ])
        
    col_widths = [6*cm, 1.5*cm] + [3*cm]*6
    t = Table(table_data, colWidths=col_widths, repeatRows=2)
    
    base_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'), # GLOBAL BOLD
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,2), (0,-1), 'LEFT'), # Products left align
        
        # Merges
        ('SPAN', (0,0), (0,1)), # Designation
        ('SPAN', (1,0), (1,1)), # Unit
        ('SPAN', (2,0), (3,0)), # Journee
        ('SPAN', (4,0), (5,0)), # Mois
        ('SPAN', (6,0), (7,0)), # Annee
        
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]
    
    # Conditional Styles (Skip 2 header rows, Skip Col 0,1 (Des, U))
    cond_styles = get_conditional_styles(table_data[2:], start_row=2, start_col=0)
    
    t.setStyle(TableStyle(base_styles + cond_styles))
    
    elements.append(t)
    elements.append(Spacer(1, 2*cm))
    
    # Footer
    footer_data = [
        ["Section Facturation", "", "Chef Service Commercial"]
    ]
    t_foot = Table(footer_data, colWidths=[8*cm, 8*cm, 8*cm])
    t_foot.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t_foot)
    
    doc.build(elements)
    return output_path

def get_conditional_styles(data_matrix, start_row=0, start_col=0):
    """
    Generate ReportLab TableStyle commands for conditional formatting.
    Positives (>0) -> Green
    Negatives (<0) -> Orange (#ff9800)
    Zeros/Text -> Blue (if 0) or Black (default)
    """
    styles = []
    orange_color = colors.HexColor('#ff9800')
    green_color = colors.green
    blue_color = colors.blue
    
    for r_idx, row in enumerate(data_matrix):
        for c_idx, cell_value in enumerate(row):
            # Val cleanup
            val_str = str(cell_value).replace(" DA", "").replace(" ", "").replace(",", ".").replace("%", "")
            try:
                # Check for empty string or non-numeric first
                if not val_str.strip():
                    continue

                val = float(val_str)
                actual_row = r_idx + start_row
                actual_col = c_idx + start_col
                
                if val > 0.001:
                    styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), green_color))
                elif val < -0.001:
                    styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), orange_color))
                else:
                    # Effectively Zero
                    styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), blue_color))
            except (ValueError, TypeError):
                # Text or other non-numeric content -> Default Black
                pass
                
    return styles

def generate_movements_valorises_pdf(date_str, output_path=None):
    from logic import get_logic
    logic = get_logic()
    result = logic.get_movements_valorises_data(date_str)
    data = result['data']
    totals = result['totals']
    
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Etat_Mouvements_Stocks_Valorises_{date_str}_{timestamp}.pdf"

    # Increased topMargin to accommodate fixed header
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=0.5*cm, leftMargin=0.5*cm,
                            topMargin=3.5*cm, bottomMargin=1.5*cm) 
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Define Header Drawing Function
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        from utils import check_logo_exists
        if check_logo_exists():
            try:
                # Draw Image directly on canvas
                # Top-Left corner validation
                logo_path = "logo_gica.png"
                img_width = 4*cm
                img_height = 2*cm
                # Position: Left margin, Top of page - margin + buffer ?? 
                # Actually, canvas (0,0) is bottom-left. 
                # Top of page is A4[1] (height).
                page_width, page_height = landscape(A4)
                
                x_pos = 0.5 * cm # Left margin
                y_pos = page_height - 0.5*cm - img_height # Top margin area
                
                canvas.drawImage(logo_path, x_pos, y_pos, width=img_width, height=img_height, preserveAspectRatio=True, mask='auto')
            except Exception as e:
                print(f"Error drawing logo: {e}")

        # 2. Title
        title_style = styles['Heading1']
        title_style.alignment = 1 # Center
        date_fmt = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        title_text = f"ETAT DES MOUVEMENTS DES STOCKS VALORISES - JOURNEE DU {date_fmt}"
        
        # Draw string centered
        canvas.setFont('Helvetica-Bold', 14)
        page_width, page_height = landscape(A4)
        text_width = canvas.stringWidth(title_text, 'Helvetica-Bold', 14)
        canvas.drawString((page_width - text_width) / 2.0, page_height - 2*cm, title_text)
        
        canvas.restoreState()

    # REMOVED: Flowable Logo and Title from 'elements'
    
    # TABLE 1: QUANTITIES
    p_style = ParagraphStyle('Bold', parent=styles['Normal'], fontName='Helvetica-Bold', alignment=1)
    elements.append(Paragraph("TABLEAU 1: QUANTITES", p_style))
    elements.append(Spacer(1, 0.2*cm))
    
    # Headers
    h1 = ["Désignation", "U", "JOURNEE", "", "", "MOIS", "", "", "ANNEE", "", "", "STOCK FINAL"]
    h2 = ["", "", "S.Init", "Entrées", "Sorties", "S.Init", "Entrées", "Sorties", "S.Init", "Entrées", "Sorties", ""]
    
    t1_data = [h1, h2]
    
    for row in data:
        t1_data.append([
            row['designation'],
            row['unite'],
            format_currency_report(row['day']['init']),
            format_currency_report(row['day']['in']),
            format_currency_report(row['day']['out']),
            format_currency_report(row['month']['init']),
            format_currency_report(row['month']['in']),
            format_currency_report(row['month']['out']),
            format_currency_report(row['year']['init']),
            format_currency_report(row['year']['in']),
            format_currency_report(row['year']['out']),
            format_currency_report(row['final'])
        ])
        
    # TOTAL ROW
    t1_data.append([
        "TOTAL", "", 
        format_currency_report(totals['day']['init']),
        format_currency_report(totals['day']['in']),
        format_currency_report(totals['day']['out']),
        format_currency_report(totals['month']['init']),
        format_currency_report(totals['month']['in']),
        format_currency_report(totals['month']['out']),
        format_currency_report(totals['year']['init']),
        format_currency_report(totals['year']['in']),
        format_currency_report(totals['year']['out']),
        format_currency_report(totals['final'])
    ])

    # Col Widths
    cw = 2.15*cm
    col_widths = [5*cm, 1.8*cm] + [cw]*10
    
    t1 = Table(t1_data, colWidths=col_widths, repeatRows=2)
    
    # Base Styles
    base_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'), # GLOBAL BOLD
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,2), (0,-2), 'LEFT'), # Products left align (exclude header/total)
        ('WORDWRAP', (0,0), (-1,-1), 'CJK'), # Allow wrapping
        
        # Merges
        ('SPAN', (0,0), (0,1)), # Des
        ('SPAN', (1,0), (1,1)), # Unit
        ('SPAN', (2,0), (4,0)), # Day
        ('SPAN', (5,0), (7,0)), # Month
        ('SPAN', (8,0), (10,0)), # Year
        ('SPAN', (11,0), (11,1)), # Final
        
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey), # Total Row Grey
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
    ]
    
    # Apply Conditional Formatting (Skip headers [0,1])
    cond_styles = get_conditional_styles(t1_data[2:], start_row=2, start_col=0) # Start col 0 to match data
    
    t1.setStyle(TableStyle(base_styles + cond_styles))
    elements.append(t1)
    elements.append(Spacer(1, 0.5*cm))
    
    # TABLE 2: VALUES
    elements.append(Paragraph("TABLEAU 2: VALEURS (DA)", p_style))
    elements.append(Spacer(1, 0.2*cm))
    
    # Same Header Structure but for values
    h1_v = ["Désignation", "Cout U.", "JOURNEE", "", "", "MOIS", "", "", "ANNEE", "", "", "VAL. FINALE"]
    h2_v = ["", "", "S.Init", "Entrées", "Sorties", "S.Init", "Entrées", "Sorties", "S.Init", "Entrées", "Sorties", ""]
    
    t2_data = [h1_v, h2_v]
    
    for row in data:
        t2_data.append([
            row['designation'],
            format_currency_report(row['cout_unitaire']),
            format_currency_report(row['values']['day']['init']),
            format_currency_report(row['values']['day']['in']),
            format_currency_report(row['values']['day']['out']),
            format_currency_report(row['values']['month']['init']),
            format_currency_report(row['values']['month']['in']),
            format_currency_report(row['values']['month']['out']),
            format_currency_report(row['values']['year']['init']),
            format_currency_report(row['values']['year']['in']),
            format_currency_report(row['values']['year']['out']),
            format_currency_report(row['val_final'])
        ])
    
    # TOTAL ROW FOR VALUES
    v_totals = [0.0] * 10 # 10 value columns
    
    for row in data:
        v_totals[0] += row['values']['day']['init']
        v_totals[1] += row['values']['day']['in']
        v_totals[2] += row['values']['day']['out']
        v_totals[3] += row['values']['month']['init']
        v_totals[4] += row['values']['month']['in']
        v_totals[5] += row['values']['month']['out']
        v_totals[6] += row['values']['year']['init']
        v_totals[7] += row['values']['year']['in']
        v_totals[8] += row['values']['year']['out']
        v_totals[9] += row['val_final']
        
    t2_data.append([
        "TOTAL", "", 
        format_currency_report(v_totals[0]),
        format_currency_report(v_totals[1]),
        format_currency_report(v_totals[2]),
        format_currency_report(v_totals[3]),
        format_currency_report(v_totals[4]),
        format_currency_report(v_totals[5]),
        format_currency_report(v_totals[6]),
        format_currency_report(v_totals[7]),
        format_currency_report(v_totals[8]),
        format_currency_report(v_totals[9])
    ])
    
    t2 = Table(t2_data, colWidths=col_widths, repeatRows=2)
    # Apply Conditional Formatting (Skip headers [0,1], Start Col 0)
    cond_styles_v = get_conditional_styles(t2_data[2:], start_row=2, start_col=0)
    
    t2.setStyle(TableStyle(base_styles + cond_styles_v)) # Reuse base style
    elements.append(t2)
    elements.append(Spacer(1, 0.5*cm))
    
    # Signature Blocks
    sig_data = [
        ["Section Facturation", "Le Chef Service Commercial", "Chef Service Comptabilité", "Le Chef Depot/Assistant PDG"],
        ["", "", "", ""] # Space for signing
    ]
    
    t_sig = Table(sig_data, colWidths=[7*cm, 7*cm, 7*cm, 7*cm])
    t_sig.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('MINROWHEIGHT', (1,0), (1,0), 2*cm), # Space for signature
    ]))
    
    elements.append(t_sig)
    
    # Use NumberedCanvas for Page X / Y AND draw_header for repeating header
    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return output_path

def generate_annual_receivables_excel(data, date_n, output_path=None):
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Etat_Creances_Annuelles_{date_n}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etat Creances Annuelles"
    
    # Styles
    bold_font = Font(bold=True, name='Arial', size=10)
    title_font = Font(bold=True, name='Arial', size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Title
    date_fmt = datetime.strptime(date_n, "%Y-%m-%d").strftime("%d/%m/%Y")
    ws['B2'] = f"ÉTAT RÉCAPITULATIF ANNUEL DES CRÉANCES ET RECOUVREMENTS CLIENTS (SITUATION AU {date_fmt})"
    ws['B2'].font = title_font
    ws['B2'].alignment = center_align
    ws.merge_cells('B2:G2')
    
    # Headers
    headers = [
        ("A4", "Raison Sociale"),
        ("B4", "Solde au 01/01"),
        ("C4", "Achats (Année)"),
        ("D4", "Paiements (Année)"),
        ("E4", "Solde Final (Jour N)"),
        ("F4", "% Recouvrement")
    ]
    
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        
    # Data
    row_idx = 5
    for row in data['data']:
        ws.cell(row=row_idx, column=1, value=row['raison_sociale']).border = border
        
        ws.cell(row=row_idx, column=2, value=format_currency_report(row['solde_01_01'])).border = border
        ws.cell(row=row_idx, column=3, value=format_currency_report(row['achats'])).border = border
        ws.cell(row=row_idx, column=4, value=format_currency_report(row['paiements'])).border = border
        ws.cell(row=row_idx, column=5, value=format_currency_report(row['solde_final'])).border = border
        
        perc_val = f"{row['recouvrement']:.1f}%"
        ws.cell(row=row_idx, column=6, value=perc_val).border = border
        ws.cell(row=row_idx, column=6).alignment = center_align
        
        row_idx += 1
        
    # Totals Row
    totals = data['totals']
    ws.cell(row=row_idx, column=1, value=f"SOLDE GLOBAL DES CRÉANCES AU {date_fmt}").font = bold_font
    ws.cell(row=row_idx, column=1).border = border
    
    ws.cell(row=row_idx, column=2, value=format_currency_report(totals['solde_init'])).font = bold_font
    ws.cell(row=row_idx, column=2).border = border
    
    ws.cell(row=row_idx, column=3, value=format_currency_report(totals['achats'])).font = bold_font
    ws.cell(row=row_idx, column=3).border = border
    
    ws.cell(row=row_idx, column=4, value=format_currency_report(totals['paiements'])).font = bold_font
    ws.cell(row=row_idx, column=4).border = border
    
    ws.cell(row=row_idx, column=5, value=format_currency_report(totals['solde_final'])).font = bold_font
    ws.cell(row=row_idx, column=5).border = border
    
    ws.cell(row=row_idx, column=6, value="").border = border
    
    # Column Widths
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
         ws.column_dimensions[col].width = 18
    ws.column_dimensions['F'].width = 15

    wb.save(output_path)
    return output_path

def generate_annual_receivables_pdf(data, date_n, output_path=None):
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Etat_Creances_Annuelles_{date_n}_{timestamp}.pdf"

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Logo Check
    from utils import check_logo_exists
    if check_logo_exists():
        try:
            from reportlab.platypus import Image as RLImage
            im = RLImage("logo_gica.png", width=4*cm, height=2*cm)
            im.hAlign = 'LEFT'
            elements.append(im)
        except: pass
        
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    date_fmt = datetime.strptime(date_n, "%Y-%m-%d").strftime("%d/%m/%Y")
    title = Paragraph(f"ÉTAT RÉCAPITULATIF ANNUEL DES CRÉANCES ET RECOUVREMENTS CLIENTS<br/>(SITUATION AU {date_fmt})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 1*cm))
    
    # Table Data
    headers = ["Raison Sociale", "Solde au 01/01", "Achats (Année)", "Paiements (Année)", "Solde Final", "% Recouvrement"]
    
    table_data = [headers]
    
    for row in data['data']:
        table_data.append([
            row['raison_sociale'],
            format_currency_report(row['solde_01_01']),
            format_currency_report(row['achats']),
            format_currency_report(row['paiements']),
            format_currency_report(row['solde_final']),
            f"{row['recouvrement']:.1f}%"
        ])
        
    # Totals Row
    totals = data['totals']
    table_data.append([
        f"SOLDE GLOBAL AU {date_fmt}",
        format_currency_report(totals['solde_init']),
        format_currency_report(totals['achats']),
        format_currency_report(totals['paiements']),
        format_currency_report(totals['solde_final']),
        ""
    ])
    
    # Column Widths
    col_widths = [7*cm, 4*cm, 4*cm, 4*cm, 4*cm, 3*cm]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Base Styles
    base_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'), # GLOBAL BOLD
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,1), (0,-1), 'LEFT'), # Raison Sociale Left
        
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), # Header BG
        ('FONTSIZE', (0,0), (-1,-1), 8),
        
        # Total Row Style
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        # ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), # Redundant
    ]
    
    # Conditional Styles (Skip 1 header row, Skip Col 0 (Raison Soc))
    cond_styles = get_conditional_styles(table_data[1:], start_row=1, start_col=0)
    
    t.setStyle(TableStyle(base_styles + cond_styles))

    
    elements.append(t)
    elements.append(Spacer(1, 2*cm))
    
    # Signature Blocks
    sig_data = [
        ["Chef de Service Commercial", "", "Service Comptabilité"]
    ]
    t_sig = Table(sig_data, colWidths=[8*cm, 5*cm, 8*cm])
    t_sig.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('MINROWHEIGHT', (1,0), (1,0), 2*cm),
    ]))
    
    elements.append(t_sig)
    
    doc.build(elements)
    return output_path

def generate_grand_livre_pdf(data, period, output_path=None):
    """
    Generate Detailed Grand Livre PDF
    """
    if output_path is None:
        if not os.path.exists("Exports_PDF"): os.makedirs("Exports_PDF")
        output_path = os.path.join("Exports_PDF", f"Grand_Livre_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    
    from reportlab.platypus import PageBreak, Image as ReportLabImage
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1, # CENTER
        spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        'CustomSubTitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        spaceAfter=20
    )
    
    # Check if data exists
    if not data:
        elements.append(Paragraph("Aucune donnée trouvée pour la période.", styles['Normal']))
        doc.build(elements)
        return output_path

    # Iterate Clients
    for idx, client_data in enumerate(data):
        if idx > 0:
            elements.append(PageBreak())
            
        client = client_data['client']
        
        # --- HEADER ---
        # Logo Logic (Simplified)
        logo_path = "logo_entete.png" if os.path.exists("logo_entete.png") else "logo.png"
        if os.path.exists(logo_path):
            im = ReportLabImage(logo_path, width=4*cm, height=2.5*cm) # Adjust aspect ratio
            # Use a Table for Header Layout to center Title and align Logo
            # Col1: Logo, Col2: Title Info
            start_fmt = datetime.strptime(period['start'], '%Y-%m-%d').strftime('%d/%m/%Y')
            end_fmt = datetime.strptime(period['end'], '%Y-%m-%d').strftime('%d/%m/%Y')
            
            title_text = "<b>GRAND-LIVRE DÉTAILLÉ DES OPÉRATIONS CLIENTS</b>"
            sub_text = f"SITUATION DU {start_fmt} AU {end_fmt}"
            
            header_table_data = [[
                im, 
                [Paragraph(title_text, title_style), Paragraph(sub_text, subtitle_style)]
            ]]
            
            t_head = Table(header_table_data, colWidths=[5*cm, 20*cm])
            t_head.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'CENTER'),
            ]))
            elements.append(t_head)
        else:
            # Text Only Header
            start_fmt = datetime.strptime(period['start'], '%Y-%m-%d').strftime('%d/%m/%Y')
            end_fmt = datetime.strptime(period['end'], '%Y-%m-%d').strftime('%d/%m/%Y')
            elements.append(Paragraph("GRAND-LIVRE DÉTAILLÉ DES OPÉRATIONS CLIENTS", title_style))
            elements.append(Paragraph(f"SITUATION DU {start_fmt} AU {end_fmt}", subtitle_style))
            
        elements.append(Spacer(1, 0.5*cm))
        
        # Client Info
        c_info = f"<b>CLIENT: {client['raison_sociale']}</b> ({client['code_client'] or 'N/A'})"
        elements.append(Paragraph(c_info, styles['Heading2']))
        elements.append(Spacer(1, 0.3*cm))
        
        # --- TABLE DATA ---
        # Cols: Date, Ref, Libellé, Debit, Credit, Solde
        # Widths: ~27.7cm available (A4 Landscape 29.7 - 2 margins)
        # Date(2.5), Ref(2.5), Lib(9.7), Deb(3.5), Cred(3.5), Solde(4) => 25.7 ok
        col_widths = [2.5*cm, 2.5*cm, 9.7*cm, 3.5*cm, 3.5*cm, 4*cm]
        headers = ["DATE", "RÉF", "LIBELLÉ", "DÉBIT", "CRÉDIT", "SOLDE"]
        
        table_rows = [headers]
        
        # Initial Balance
        s_date = datetime.strptime(period['start'], '%Y-%m-%d').strftime('%d/%m/%Y')
        init_bal = client_data['initial_balance']
        table_rows.append([
            s_date, "-", "SOLDE INITIAL", "", "", format_currency_report(init_bal)
        ])
        
        # Movements
        for mv in client_data['movements']:
            # Formatting
            try: d_str = datetime.strptime(mv['date'], '%Y-%m-%d').strftime('%d/%m/%Y')
            except: d_str = mv['date']
            
            ref = str(mv['ref'])
            lib = mv['libelle']
            deb = format_currency_report(mv['debit']) if mv['debit'] != 0 else "-"
            cred = format_currency_report(mv['credit']) if mv['credit'] != 0 else "-"
            solde = format_currency_report(mv['solde_progressif'])
            
            table_rows.append([d_str, ref, lib, deb, cred, solde])
            
        # Totals
        table_rows.append([
            "", "", "TOTAUX PÉRIODE", 
            format_currency_report(client_data['total_debit']), 
            format_currency_report(client_data['total_credit']), 
            ""
        ])
        
        # Final
        e_date = datetime.strptime(period['end'], '%Y-%m-%d').strftime('%d/%m/%Y')
        fin_bal = client_data['final_balance']
        table_rows.append([
            "", "", f"SOLDE FINAL AU {e_date}", 
            "", "", format_currency_report(fin_bal)
        ])
        
        # Create Table
        t = Table(table_rows, colWidths=col_widths, repeatRows=1)
        
        # Styles
        tbl_styles = [
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), # Header Bold
            ('ALIGN', (0,0), (-1,0), 'CENTER'), # Header Align
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), # Header BG
            
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (1,-1), 'CENTER'), # Date/Ref Center
            ('ALIGN', (3,0), (-1,-1), 'RIGHT'), # Numbers Right
            
            # Initial Solde Row (Row 1)
            ('BACKGROUND', (2,1), (2,1), colors.whitesmoke),
            ('FONTNAME', (2,1), (2,1), 'Helvetica-Bold'),
            
            # Totals Row (Second to last)
            ('BACKGROUND', (2,-2), (4,-2), colors.lightgrey),
            ('FONTNAME', (2,-2), (4,-2), 'Helvetica-Bold'),
            
            # Final Row (Last)
            ('BACKGROUND', (2,-1), (2,-1), colors.wheat), # Highligh Label
            ('BACKGROUND', (5,-1), (5,-1), colors.wheat), # Highlight Value
            ('FONTNAME', (2,-1), (-1,-1), 'Helvetica-Bold'),
        ]
        
        # Conditional Formatting Loop
        # Row 0 is Header. 
        # Row 1 is Init.
        # Rows 2 to (-2) are movements.
        
        # We need to map table row index to movement index to check 'is_cancelled'
        # Table Row i corresponds to movement i-2 (since row 0=Head, row 1=Init)
        # Length of movements = len(table_rows) - 3 (Head, Init, Totals, Final ... wait 4 rows extra?)
        # Let's count properly:
        # Rows: Head(0) -> Init(1) -> Moves(2...N) -> Total(N+1) -> Final(N+2)
        
        start_moves = 2
        moves = client_data['movements']
        
        for i, mv in enumerate(moves):
            row_idx = start_moves + i
            if mv.get('is_cancelled'):
                # Orange Text
                tbl_styles.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.orange))
                tbl_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                
            # Solde Color (Positive = Advance = Blue, Negative = Debt = Black)
            # Solde is Col 5
            try:
                val = float(mv['solde_progressif'])
                if val > 0.001:
                    tbl_styles.append(('TEXTCOLOR', (5, row_idx), (5, row_idx), colors.blue))
                else:
                    # Debt or Zero
                    if not mv.get('is_cancelled'):
                        tbl_styles.append(('TEXTCOLOR', (5, row_idx), (5, row_idx), colors.black))
            except: pass
            
        # Final Row Solde Color
        try:
            val = float(fin_bal)
            if val > 0.001: # Advance
                tbl_styles.append(('TEXTCOLOR', (5, -1), (5, -1), colors.blue))
            else: # Debt
                tbl_styles.append(('TEXTCOLOR', (5, -1), (5, -1), colors.black))
        except: pass

        t.setStyle(TableStyle(tbl_styles))
        elements.append(t)
        
    doc.build(elements, canvasmaker=NumberedCanvas)
    return output_path

def generate_recovery_pdf(data, month, year, output_path):
    """
    Generate PDF for Suivi Recouvrement Mensuel (M-1)
    """
    if not data or not data.get('data'):
        return None

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
                            
    elements = []
    styles = getSampleStyleSheet()
    
    # --- HEADER ---
    # Logo
    from utils import check_logo_exists
    if check_logo_exists():
        try:
            from reportlab.platypus import Image as RLImage
            im = RLImage("logo_gica.png", width=4*cm, height=2*cm)
            im.hAlign = 'LEFT'
            elements.append(im)
        except: pass
        
    elements.append(Spacer(1, 0.5*cm))
    
    # Title
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    title = Paragraph(f"ÉTAT DE COUVERTURE DES CRÉANCES (MOIS {month:02d}/{year})", title_style)
    elements.append(title)
    
    # Global Stats
    if data.get('totals'):
        rate = data['totals'].get('rate', 0.0)
        # Color logic
        color = "green" if rate >= 80 else "orange" if rate > 50 else "red"
        rate_text = f'<font color="{color}">Taux de Recouvrement Global : {rate:.2f}%</font>'
        
        stat_style = ParagraphStyle('StatStyle', parent=styles['Normal'], fontSize=12, alignment=1, spaceBefore=6)
        elements.append(Paragraph(rate_text, stat_style))
        
    elements.append(Spacer(1, 1*cm))
    
    # --- TABLE ---
    headers = ["Client", "Dette M-1\n(Cible)", "Paiements M\n(Réalisé)", "Reste à Payer\n(Ecart)", "Statut"]
    
    table_data = [headers]
    
    for row in data['data']:
        table_data.append([
            row['raison_sociale'],
            format_currency_report(row['dette_m_1']),
            format_currency_report(row['paiements_m']),
            format_currency_report(row['reste_a_payer']),
            row['statut']
        ])
        
    # Totals Row
    if data.get('totals'):
        t = data['totals']
        table_data.append([
            "TOTAL",
            format_currency_report(t['target']),
            format_currency_report(t['realized']),
            "",  # No total for gap usually, or maybe yes? Let's leave empty as interface.
            ""
        ])

    # Column Widths
    # A4 Width = ~21cm. Margins 2cm total. Avail ~19cm.
    # Client (8), Dette (3), Pay (3), Reste (3), Statut (3) -> 20cm too wide?
    # Let's try: Client 7, Numbers 3 each (9), Statut 3 -> 19cm.
    col_widths = [6*cm, 3.25*cm, 3.25*cm, 3.25*cm, 3.25*cm]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Styles
    msg_styles = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,1), (3,-1), 'RIGHT'), # Numbers Right
        ('ALIGN', (4,1), (4,-1), 'CENTER'), # Status Center
        
        # Total Row
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
    ]
    
    # Conditional Formatting for rows
    # Row 0 is header. data row i corresponds to table row i+1
    for i, row in enumerate(data['data']):
        row_idx = i + 1
        statut = row['statut']
        status_col_idx = 4
        
        if statut == "RÉGLÉ":
            # Text Green
            msg_styles.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.green))
        elif statut == "ALERTE RECOUVREMENT":
            # Text Red
            msg_styles.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.red))
            msg_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
        else:
            # En Attente -> Orange
            msg_styles.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.orange))
            
    t.setStyle(TableStyle(msg_styles))
    elements.append(t)
    
    elements.append(Spacer(1, 2*cm))
    
    # --- FOOTER ---
    # Signature "Service Recouvrement"
    footer_data = [["Service Recouvrement", "", "Direction Commerciale"]]
    t_foot = Table(footer_data, colWidths=[6*cm, 7*cm, 6*cm])
    t_foot.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t_foot)


    doc.build(elements, canvasmaker=NumberedCanvas)
    return output_path


def generate_pareto_pdf(data: Dict[str, Any], start_date: str, end_date: str, filename: str):
    """
    Generate Pareto Analysis PDF with Charts and Table.
    """
    from utils import check_logo_exists, format_currency, generate_pareto_charts
    
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    story = []
    width, height = A4
    styles = getSampleStyleSheet()
    
    # 1. Header
    logo_path = check_logo_exists()
    title = "ANALYSE DE PERFORMANCE COMMERCIALE (PARETO)"
    
    title_style = ParagraphStyle(
        'ParetoTitle', parent=styles['Heading1'], fontSize=16, 
        textColor=colors.HexColor('#1a237e'), alignment=TA_CENTER, spaceAfter=20
    )
    
    if logo_path:
        img = Image(logo_path, width=2.5*cm, height=2.5*cm)
        img.hAlign = 'LEFT'
        header_data = [[img, Paragraph(title, title_style)]]
        header_table = Table(header_data, colWidths=[3*cm, 15*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(header_table)
    else:
        story.append(Paragraph(title, title_style))
        
    story.append(Paragraph(f"<b>Période du :</b> {start_date} <b>au</b> {end_date}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # 2. Generate and Insert Charts
    # -----------------------------
    curve_path, pie_path = generate_pareto_charts(data['data'])
    
    # Legend Styles
    legend_style = ParagraphStyle(
        'LegendStyle', parent=styles['Normal'], fontSize=9, textColor=colors.black, 
        alignment=TA_CENTER, spaceBefore=0, spaceAfter=0
    )
    legend_title_style = ParagraphStyle(
        'LegendTitle', parent=styles['Normal'], fontSize=9, textColor=colors.black, 
        fontName='Helvetica-Bold', alignment=TA_CENTER, spaceBefore=6, spaceAfter=2
    )

    if curve_path and os.path.exists(curve_path):
        img_curve = Image(curve_path, width=16*cm, height=9.6*cm)
        story.append(img_curve)
        
        # Pareto Legend
        story.append(Paragraph("Interprétation du Diagramme :", legend_title_style))
        story.append(Paragraph("Les barres classent vos clients par volume de chiffre d'affaires. La ligne rouge montre le cumul. La zone où la ligne coupe la barre des 80% identifie vos clients stratégiques (Classe A).", legend_style))
        story.append(Spacer(1, 0.5*cm))
        
    if pie_path and os.path.exists(pie_path):
        img_pie = Image(pie_path, width=10*cm, height=10*cm)
        story.append(img_pie)
        
        # ABC Legend
        story.append(Spacer(1, 0.2*cm))
        
        # A
        story.append(Paragraph('<font color="#2e7d32"><b>Classe A (Vert) :</b></font> Représente 80% de votre activité. Ce sont vos clients piliers. Toute baisse de leur part est un risque majeur pour l\'unité.', legend_style))
        # B
        story.append(Paragraph('<font color="#1565c0"><b>Classe B (Bleu) :</b></font> Représente les 15% suivants. Ce sont vos clients en développement ou à fort potentiel.', legend_style))
        # C
        story.append(Paragraph('<font color="#757575"><b>Classe C (Gris) :</b></font> Représente les derniers 5%. Ce sont des clients occasionnels qui génèrent un grand volume administratif pour un faible revenu.', legend_style))
        
        story.append(Spacer(1, 0.8*cm))
        
    # 3. Dynamic Synthesis
    # --------------------
    clients_a = 0
    total_clients = len(data['data'])
    for row in data['data']:
        if row['classe'] == 'A': clients_a += 1
        
    if total_clients > 0:
        perc_clients_a = (clients_a / total_clients) * 100
        
        # Risk Warning
        warning_text = ""
        if clients_a < 5 or perc_clients_a < 10: # "Petit groupe" definition
             warning_text = " <b>Attention à la forte dépendance envers ce petit groupe.</b>"
             
        synthesis_text = f"<b>Analyse de l'unité :</b> Votre activité est concentrée sur <b>{clients_a}</b> clients qui réalisent à eux seuls 80% du chiffre d'affaires.{warning_text}"
        
        synth_style = ParagraphStyle(
            'SynthStyle', parent=styles['Normal'], fontSize=11, 
            textColor=colors.HexColor('#1a237e'), alignment=TA_CENTER, 
            backColor=colors.HexColor('#e8eaf6'), borderWidth=1, borderColor=colors.HexColor('#1a237e'),
            borderPadding=10, spaceAfter=20
        )
        story.append(Paragraph(synthesis_text, synth_style))
        story.append(Spacer(1, 0.5*cm))

    # 4. Table
    # --------
    # Columns: Rang | Client | CA (DA) | % Cumulé | Classe
    table_headers = ['Rang', 'Client', 'Chiffre d\'Affaires', '% Cumulé', 'Classe']
    table_data = [table_headers]
    
    for row in data['data']:
        table_data.append([
            str(row['rank']),
            row['client_name'],
            format_currency(row['ca']),
            f"{row['cumul_perc']:.2f}%",
            row['classe']
        ])
        
    col_widths = [1.5*cm, 8*cm, 4*cm, 2.5*cm, 2*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Styles
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('ALIGN', (1,1), (1,-1), 'LEFT'), # Client Name
        ('ALIGN', (2,1), (2,-1), 'RIGHT'), # CA
    ]
    
    for i, row in enumerate(data['data']):
        row_idx = i + 1
        if row['classe'] == 'A':
            t_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#e8f5e9')))
            t_style.append(('TEXTCOLOR', (4, row_idx), (4, row_idx), colors.green))
        elif row['classe'] == 'B':
            t_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#e3f2fd')))
            t_style.append(('TEXTCOLOR', (4, row_idx), (4, row_idx), colors.blue))
    
    t.setStyle(TableStyle(t_style))
    story.append(t)
    
    # Footer
    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.drawString(1*cm, 0.75*cm, "Généré par le Système de Gestion Commerciale")
        canvas.drawRightString(A4[0]-1*cm, 0.75*cm, f"Page {doc.page}")
        canvas.restoreState()
        
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return filename

def generate_etat_104_pdf(data: list, start_date: str, end_date: str, output_path: str):
    """
    Générer PDF de l'État 104 (Ventes par client).
    
    Format standard avec:
    - En-tête avec logo GICA
    - Titre : "ETAT 104 DE L'ANNÉE : [année]"
    - Tableau avec totaux
    - Numérotation automatique
    """
    from datetime import datetime
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    import os
    
    # Create PDF with custom canvas for headers
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=1.5*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # En-tête avec logo
    logo_path = "LOGO GICA.png"
    if not os.path.exists(logo_path):
        logo_path = "logo_entete.png"
    
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=2*cm, height=2*cm)
            elements.append(logo)
            elements.append(Spacer(1, 0.3*cm))
        except:
            pass
    
    # Titre principal
    year = datetime.strptime(start_date, "%Y-%m-%d").year
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a237e'),
        alignment=TA_CENTER,
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    title = Paragraph(f"<b>ÉTAT 104 DE L'ANNÉE : {year}</b>", title_style)
    elements.append(title)
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    subtitle = Paragraph(f"Période : {start_date} au {end_date}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.5*cm))
    
    # Tableau de données
    table_data = [[
        "N° d'ordre",
        "Raison Sociale / Nom du Client",
        "Adresse précise",
        "NIF\n(15 chiffres)",
        "Article\nd'imposition\n(A)",
        "Montant des\nventes (HT)",
        "Montant de\nla TVA",
        "Montant TTC"
    ]]
    
    total_ht = 0.0
    total_tva = 0.0
    total_ttc = 0.0
    
    for row in data:
        table_data.append([
            str(row['numero']),
            row['raison_sociale'],
            row['adresse'],
            row['nif'],
            row['article_imposition'],
            format_currency_report(row['total_ht']),
            format_currency_report(row['total_tva']),
            format_currency_report(row['total_ttc'])
        ])
        total_ht += row['total_ht']
        total_tva += row['total_tva']
        total_ttc += row['total_ttc']
    
    # Ligne de total
    table_data.append([
        "",
        "",
        "",
        "",
        "TOTAL",
        format_currency_report(total_ht),
        format_currency_report(total_tva),
        format_currency_report(total_ttc)
    ])
    
    # Créer le tableau
    table = Table(table_data, colWidths=[1.5*cm, 5*cm, 4.5*cm, 3*cm, 2.5*cm, 3*cm, 3*cm, 3*cm])
    
    # Style du tableau
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Corps du tableau
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # N° centré
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),  # NIF et Article centrés
        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),  # Montants à droite
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1a237e')),
        ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
        ('ALIGN', (5, -1), (-1, -1), 'RIGHT'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"✅ PDF État 104 généré : {output_path}")
    return output_path


def generate_cockpit_pdf(data: Dict[str, Any]):

    """
    Generate Master Dashboard (Cockpit) PDF (One-Page Landscape).
    """
    from utils import ensure_pdf_export_dir, check_logo_exists
    directory = ensure_pdf_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(directory, f"Cockpit_{timestamp}.pdf")
    
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4),
                          rightMargin=1*cm, leftMargin=1*cm,
                          topMargin=1*cm, bottomMargin=1*cm)
    story = []
    styles = getSampleStyleSheet()
    
    # Header
    # ------
    logo_path = check_logo_exists()
    header_data = [[
        Image(logo_path, width=3*cm, height=2*cm) if logo_path else "",
        Paragraph("<b>TABLEAU DE BORD MAÎTRE (COCKPIT)</b>", styles['Title']),
        Paragraph(f"<b>Période: {data.get('period', '--')}</b>", styles['Normal'])
    ]]
    t_head = Table(header_data, colWidths=[4*cm, 18*cm, 5*cm])
    t_head.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t_head)
    story.append(Spacer(1, 1*cm))
    
    # 1. KPI Tiles
    # ------------
    kpis = data['kpis']
    
    # Helper to create Tile content
    def create_tile_content(title, value, sub, bg_color):
        s_title = ParagraphStyle('TTitle', parent=styles['Normal'], fontSize=10, textColor=colors.white, fontName='Helvetica-Bold')
        s_val = ParagraphStyle('TVal', parent=styles['Normal'], fontSize=20, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
        s_sub = ParagraphStyle('TSub', parent=styles['Normal'], fontSize=9, textColor=colors.white, alignment=TA_CENTER)
        
        return [
            Paragraph(title, s_title),
            Spacer(1, 0.4*cm),
            Paragraph(value, s_val),
            Spacer(1, 0.2*cm),
            Paragraph(str(sub), s_sub)
        ]
        
    # Tiles Data
    evo = kpis['evolution']
    arrow = "▲" if evo >= 0 else "▼"
    
    tiles_row = [
        create_tile_content("PERFORMANCE VENTES", format_currency(kpis['ca_curr']), f"{arrow} {abs(evo):.1f}% vs M-1", colors.HexColor("#1e88e5")),
        create_tile_content("SANTÉ FINANCIÈRE", f"{kpis['recovery_rate']:.1f}%", "Recouvrement", colors.HexColor("#43a047")),
        create_tile_content("RISQUE CRÉANCE (+30J)", format_currency(kpis['debt_30_days']), "Montant à Risque", colors.HexColor("#fb8c00")),
        create_tile_content("ALERTE OPÉRATIONNELLE", f"{kpis['cancel_rate']:.1f}%", "Taux Annulation", colors.HexColor("#e53935")),
    ]
    
    # Table for Tiles
    # We want 4 cells with background color.
    # We can't put listing in cell easily with Flowables unless we use a nested table or just simple text.
    # Let's use nested tables for tiles to carry background
    
    tile_tables = []
    for content, color in zip(tiles_row, ["#1e88e5", "#43a047", "#fb8c00", "#e53935"]):
        t = Table([[c] for c in content], colWidths=[6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(color)),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX', (0,0), (-1,-1), 0, colors.white), # Borderless?
        ]))
        tile_tables.append(t)
        
    main_tile_table = Table([tile_tables], colWidths=[6.5*cm]*4)
    story.append(main_tile_table)
    story.append(Spacer(1, 1*cm))
    
    # 2. Charts
    # ---------
    from utils import generate_cockpit_charts
    p_a, p_b, p_c = generate_cockpit_charts(data)
    
    charts_row = []
    
    if p_a and os.path.exists(p_a):
        charts_row.append(Image(p_a, width=5*cm, height=8*cm))
    else:
        charts_row.append("No Data")
        
    if p_b and os.path.exists(p_b):
        charts_row.append(Image(p_b, width=8*cm, height=6*cm))
    else:
        charts_row.append("No Data")
        
    if p_c and os.path.exists(p_c):
        charts_row.append(Image(p_c, width=10*cm, height=6*cm))
    else:
        charts_row.append("No Data")
        
    t_charts = Table([charts_row], colWidths=[6*cm, 9*cm, 11*cm])
    t_charts.setStyle(TableStyle([
         ('ALIGN', (0,0), (-1,-1), 'CENTER'),
         ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t_charts)
    story.append(Spacer(1, 1*cm))
    
    # 3. Alerts
    # ---------
    story.append(Paragraph("<b>ALERTES DE SÉCURITÉ (Code Rouge)</b>", ParagraphStyle('HAlert', parent=styles['Heading3'], textColor=colors.red)))
    story.append(Spacer(1, 0.2*cm))
    
    alert_headers = ['Client', 'Montant à Risque', 'Motif']
    alert_data = [alert_headers]
    for alert in data['alerts']:
        alert_data.append([
            alert['name'],
            format_currency(alert['amount']),
            alert['reason']
        ])
    
    if len(alert_data) > 1:
        t_alerts = Table(alert_data, colWidths=[12*cm, 6*cm, 8*cm])
        t_alerts.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ffebee')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.red),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.red),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
            ('ALIGN', (1,1), (1,-1), 'RIGHT'),
        ]))
        story.append(t_alerts)
    else:
        story.append(Paragraph("Aucune alerte critique détectée.", styles['Normal']))
        
    # Build
    doc.build(story)
    return filename



def generate_invoice_excel(invoice_data, output_path=None):
    """
    Générer Facture au format Excel - EXACTEMENT identique à l'image de référence
    Optimisé pour:
    - Consultation à l'écran
    - Impression directe sur imprimante matricielle 80 colonnes
    - Alignement parfait avec papier préimprimé
    """
    if not output_path:
        directory = "Exports_Excel"
        if not os.path.exists(directory):
            os.makedirs(directory)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        clean_name = "".join([c for c in invoice_data['raison_sociale'] if c.isalnum() or c in (' ', '_')]).strip()
        output_path = os.path.join(directory, f"Facture_{invoice_data['numero']}_{clean_name}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"
    
    # ========================================
    # CONFIGURATION IMPRIMANTE MATRICIELLE 80 COLONNES
    # ========================================
    
    # Largeurs de colonnes optimisées (en caractères Excel)
    # Total ≈ 80 colonnes pour imprimante matricielle
    colonnes_width = {
        'A': 3,   'B': 8,   'C': 8,   'D': 25,  'E': 3,
        'F': 5,   'G': 8,   'H': 10,  'I': 12,  'J': 8,
        'K': 15,  'L': 5,   'M': 5,   'N': 5,   'O': 12,
        'P': 5,   'Q': 5,   'R': 5,   'S': 5,   'T': 5,
        'U': 5,   'V': 5,   'W': 5,   'X': 20,  'Y': 5,
        'Z': 5
    }
    
    for col, width in colonnes_width.items():
        ws.column_dimensions[col].width = width
    
    # ========================================
    # STYLES
    # ========================================
    
    font_normal = Font(name='Cambria', size=10)
    font_bold = Font(name='Cambria', size=10, bold=True)
    font_bold_11 = Font(name='Cambria', size=11, bold=True)
    font_red_bold = Font(name='Cambria', size=11, bold=True, color='FF0000')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ========================================
    # ZONE EN-TÊTE (LIGNES 1-5)
    # ========================================
    
    # Ligne 3: N° FACTURE (Colonne O-P) - En ROUGE selon image
    ws['O3'] = invoice_data.get('numero', '')
    ws['O3'].font = font_red_bold
    ws['O3'].alignment = align_center
    
    # Ligne 5: Date avec jour de la semaine en français
    try:
        dt = datetime.strptime(invoice_data['date_facture'], '%Y-%m-%d')
        # Formatage date française
        try:
            import locale
            try:
                locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            except:
                try:
                    locale.setlocale(locale.LC_TIME, 'French_France.1252')
                except:
                    locale.setlocale(locale.LC_TIME, 'fra_fra')
        except:
            pass
        
        try:
            date_with_day = dt.strftime('%A %d %B %Y')
        except:
            # Fallback manuel
            mois_fr = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                      'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
            jours_fr = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
            jour_semaine = jours_fr[dt.weekday()]
            date_with_day = f"{jour_semaine} {dt.day} {mois_fr[dt.month-1]} {dt.year}"
    except:
        date_with_day = invoice_data['date_facture']
    
    ws['O5'] = date_with_day
    ws['O5'].font = font_normal
    ws['O5'].alignment = align_left
    
    # ========================================
    # ZONE CLIENT GAUCHE (LIGNES 6-12)
    # ========================================
    
    # Ligne 6-7: Catégorie + Code Client (ex: "SPA  0216EPN4")
    categorie = invoice_data.get('client_categorie', invoice_data.get('categorie', ''))
    code_client = invoice_data.get('code_client', '')
    
    if categorie or code_client:
        client_cat_code = f"{categorie}  {code_client}".strip()
        ws['D7'] = client_cat_code
        ws['D7'].font = font_bold
        ws['D7'].alignment = align_left
    
    # Ligne 9: Raison Sociale (ex: "INFRARAIL EPE SPA")
    ws['D9'] = invoice_data.get('raison_sociale', '')
    ws['D9'].font = font_bold
    ws['D9'].alignment = align_left
    
    # Ligne 10: Adresse
    ws['D10'] = invoice_data.get('adresse', '')
    ws['D10'].font = font_bold
    ws['D10'].alignment = align_left
    
    # Ligne 11: NIF + Article Imposition
    nif = invoice_data.get('nif', '')
    article_imp = invoice_data.get('article_imposition', '')
    
    if nif or article_imp:
        ws['D11'] = f"NIF {nif}   {article_imp}".strip()
        ws['D11'].font = font_bold
        ws['D11'].alignment = align_left
    
    # ========================================
    # ZONE CONVENTION DROITE (LIGNE 8)
    # ========================================
    
    # Ligne 8 colonne X: Convention (ex: "convention n°02/2025")
    convention = invoice_data.get('convention', invoice_data.get('contrat_code', ''))
    if convention:
        ws['X8'] = f"convention n°{convention}"
        ws['X8'].font = font_normal
        ws['X8'].alignment = align_left
    
    # ========================================
    # ZONE PAIEMENT DROITE (LIGNES 10-13)
    # ========================================
    
    ws['K10'] = "Virement"
    ws['K10'].font = font_bold
    ws['K10'].alignment = align_left
    
    ws['K11'] = "BNA OUED SMAR"
    ws['K11'].font = font_bold
    ws['K11'].alignment = align_left
    
    ws['K12'] = "00 0018443"
    ws['K12'].font = font_bold
    ws['K12'].alignment = align_left
    
    ws['K13'] = "CMP/ prest n° 0"
    ws['K13'].font = font_bold
    ws['K13'].alignment = align_left
    
    # ========================================
    # ZONE PRODUITS (LIGNE 17+)
    # ========================================
    
    row_produit = 17
    
    for ligne in invoice_data['lignes']:
        # Colonne A-C: Code produit ou début nom (ex: "CAL-SPC")
        product_code = ligne.get('code_produit', '')
        product_nom = ligne.get('product_nom', '')
        
        # Si pas de code séparé, prendre début du nom
        if product_code:
            ws[f'A{row_produit}'] = product_code
        else:
            ws[f'A{row_produit}'] = product_nom[:10]
        ws[f'A{row_produit}'].font = font_bold
        ws[f'A{row_produit}'].alignment = align_left
        
        # Colonne D-F: Description produit (ex: "CEMBRAS A 12.5 N Prem G")
        ws[f'D{row_produit}'] = product_nom
        ws[f'D{row_produit}'].font = font_bold
        ws[f'D{row_produit}'].alignment = align_left
        
        # Colonne G: Unité (ex: "Tonne")
        ws[f'G{row_produit}'] = ligne.get('unite', '')
        ws[f'G{row_produit}'].font = font_bold
        ws[f'G{row_produit}'].alignment = align_center
        
        # Colonne H: Quantité (ex: "40")
        qte = ligne.get('quantite', 0)
        ws[f'H{row_produit}'] = f"{qte:,.0f}".replace(",", " ")
        ws[f'H{row_produit}'].font = font_bold
        ws[f'H{row_produit}'].alignment = align_center
        
        # Colonne I: Prix Unitaire (ex: "6,214.00")
        pu = ligne.get('prix_unitaire', 0.0)
        ws[f'I{row_produit}'] = f"{pu:,.2f}".replace(",", " ").replace(".", ",")
        ws[f'I{row_produit}'].font = font_bold
        ws[f'I{row_produit}'].alignment = align_right
        
        # Colonne J-K: Montant (ex: "248,560.00")
        montant = ligne.get('montant', 0.0)
        ws[f'J{row_produit}'] = f"{montant:,.2f}".replace(",", " ").replace(".", ",")
        ws[f'J{row_produit}'].font = font_bold
        ws[f'J{row_produit}'].alignment = align_right
        
        row_produit += 1
    
    # ========================================
    # ZONE REMISE (LIGNE 19 SI APPLICABLE)
    # ========================================
    
    row_remise = row_produit + 1
    
    # Vérifier si remise
    has_remise = invoice_data.get('montant_remise', 0) > 0 or any(l.get('taux_remise', 0) > 0 for l in invoice_data['lignes'])
    
    if has_remise:
        ws[f'D{row_remise}'] = "Remise"
        ws[f'D{row_remise}'].font = font_bold
        ws[f'D{row_remise}'].alignment = align_left
        
        # Calculer remise totale
        total_remise = sum(l.get('montant', 0) * l.get('taux_remise', 0) / 100 for l in invoice_data['lignes'])
        
        ws[f'H{row_remise}'] = "0%"
        ws[f'H{row_remise}'].font = font_bold
        ws[f'H{row_remise}'].alignment = align_center
        
        ws[f'J{row_remise}'] = f"{total_remise:,.2f}".replace(",", " ").replace(".", ",")
        ws[f'J{row_remise}'].font = font_bold
        ws[f'J{row_remise}'].alignment = align_right
        
        row_produit = row_remise + 1
    
    # ========================================
    # ZONE TOTAUX (LIGNES 22-24)
    # ========================================
    
    row_total = row_produit + 2
    
    # Ligne 22: Montant en lettres (gauche) + Total HT (droite)
    from utils import nombre_en_lettres
    montant_lettres = nombre_en_lettres(invoice_data['montant_ttc'])
    
    ws[f'B{row_total}'] = montant_lettres
    ws[f'B{row_total}'].font = font_bold
    ws[f'B{row_total}'].alignment = align_left
    
    ws[f'J{row_total}'] = f"{invoice_data['montant_ht']:,.2f}".replace(",", " ").replace(".", ",")
    ws[f'J{row_total}'].font = font_bold
    ws[f'J{row_total}'].alignment = align_right
    
    # Ligne 23: TVA 19%
    row_total += 1
    ws[f'J{row_total}'] = f"{invoice_data['montant_tva']:,.2f}".replace(",", " ").replace(".", ",")
    ws[f'J{row_total}'].font = font_bold
    ws[f'J{row_total}'].alignment = align_right
    
    # Ligne 24: Total TTC
    row_total += 1
    ws[f'J{row_total}'] = f"{invoice_data['montant_ttc']:,.2f}".replace(",", " ").replace(".", ",")
    ws[f'J{row_total}'].font = font_bold
    ws[f'J{row_total}'].alignment = align_right
    
    # ========================================
    # ZONE TRANSPORT (LIGNES 26-27)
    # ========================================
    
    row_transport = row_total + 2
    
    # Ligne 26: Chauffeur (colonne B)
    chauffeur = invoice_data.get('chauffeur', '')
    if chauffeur:
        ws[f'B{row_transport}'] = f"Chauffeur : {chauffeur}"
        ws[f'B{row_transport}'].font = font_bold
        ws[f'B{row_transport}'].alignment = align_left
    
    # Ligne 27: Matricule (colonne G combinaison tracteur/remorque)
    matricule_tracteur = invoice_data.get('matricule_tracteur', '')
    matricule_remorque = invoice_data.get('matricule_remorque', '')
    
    if matricule_tracteur or matricule_remorque:
        # Format: "06703-313-027/007332-013-03"
        if matricule_tracteur and matricule_remorque:
            matricule_complet = f"{matricule_tracteur}/{matricule_remorque}"
        else:
            matricule_complet = matricule_tracteur or matricule_remorque
            
        ws[f'G{row_transport}'] = f"Matricule : {matricule_complet}"
        ws[f'G{row_transport}'].font = font_bold
        ws[f'G{row_transport}'].alignment = align_left
    
    # ========================================
    # ZONE SIGNATURE (LIGNE 29)
    # ========================================
    
    row_signature = row_transport + 2
    
    ws[f'B{row_signature}'] = "HAMMICHE MAKHLOUF"
    ws[f'B{row_signature}'].font = font_bold
    ws[f'B{row_signature}'].alignment = align_left
    
    # ========================================
    # CONFIGURATION IMPRESSION MATRICIELLE
    # ========================================
    
    # Page Setup
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Auto
    
    # Marges réduites pour imprimante matricielle
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1
    
    # Print Options
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.print_options.gridLines = False
    
    # Masquer les lignes de grille à l'écran
    ws.sheet_view.showGridLines = False
    
    # Zone d'impression (ajuster selon besoins)
    ws.print_area = f'A1:Z{row_signature + 5}'
    
    # ========================================
    # SAUVEGARDE
    # ========================================
    
    wb.save(output_path)
    return output_path


def generate_invoice_pdf_matricielle(invoice_data, output_path=None):
    """
    Générer Facture au format PDF - Optimisé pour Imprimante Matricielle 80 colonnes
    Positionnement absolu pour éviter les problèmes de troncature
    Format identique à l'image de référence
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm, mm
    from reportlab.lib.colors import HexColor, black, red
    
    if not output_path:
        directory = "Exports_PDF"
        if not os.path.exists(directory):
            os.makedirs(directory)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        clean_name = "".join([c for c in invoice_data['raison_sociale'] if c.isalnum() or c in (' ', '_')]).strip()
        output_path = os.path.join(directory, f"Facture_Matricielle_{invoice_data['numero']}_{clean_name}_{timestamp}.pdf")
    
    # Créer le canvas PDF
    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4  # 595 x 842 points (210mm x 297mm)
    
    # ========================================
    # CONFIGURATION POLICE
    # ========================================
    
    # Police principale
    font_normal = "Helvetica"
    font_bold = "Helvetica-Bold"
    size_normal = 9
    size_title = 11
    size_small = 8
    
    # ========================================
    # POSITIONS ABSOLUES - BASÉES SUR PRÉIMPRIMÉ RÉEL
    # Système Y inversé : 0 en BAS de page
    # Mesures depuis analyse professionnelle
    # ========================================
    
    # Marges physiques imprimante matricielle
    margin_left = 15 * mm  # Après perforations
    margin_right = 15 * mm
    
    # Positions Y (depuis le BAS de la page)
    # Conversion : Y_PDF = 297mm - Y_depuis_haut
    
    # Zone N° Facture et Date (ligne ~24mm depuis haut = 273mm depuis bas)
    y_facture_num = 273 * mm  # N° Facture ROUGE
    y_date = 267 * mm         # Date complète
    
    # Zone GICA (Informations fixes - lignes 36-40mm depuis haut)
    y_gica_ligne1 = 261 * mm  # 297 - 36 = 261
    y_gica_ligne2 = 257 * mm  # 297 - 40 = 257
    
    # Zone Code Client (ligne 44mm depuis haut = 253mm depuis bas)
    y_client_cat = 253 * mm
    
    # Cadre CLIENT (lignes 50-80mm depuis haut)
    y_client_nom = 241 * mm      # 297 - 56 = 241
    y_client_adresse = 232 * mm  # 297 - 65 = 232
    y_client_rc = 227 * mm       # 297 - 70 = 227
    y_client_nis = 222 * mm      # 297 - 75 = 222
    y_client_nif = 217 * mm      # 297 - 80 = 217 (NIF + Article)
    y_client_date_enlev = 212 * mm  # 297 - 85 = 212
    
    # Cadre COMMANDE/PAIEMENT (droite, lignes 50-80mm)
    y_paiement_reglement = 233 * mm  # 297 - 64 = 233
    y_paiement_banque = 228 * mm     # 297 - 69 = 228
    y_paiement_compte = 223 * mm     # 297 - 74 = 223
    
    # Zone Convention (si existe)
    y_convention = 253 * mm  # Même ligne que catégorie
    x_convention = 125 * mm  # Cadre droit
    
    # Zone TABLE PRODUITS (début ligne 95mm depuis haut = 202mm depuis bas)
    y_produit_start = 202 * mm
    line_height = 5 * mm
    
    # Colonnes produits (positions X exactes selon préimprimé)
    x_code = 16 * mm          # Colonne DESIGNATION début
    x_description = 42 * mm   # Description complète (déplacé vers droite pour éviter chevauchement)
    x_unite = 119 * mm        # Colonne U M (centré)
    x_quantite = 138 * mm     # Colonne QUANTITE (centré)
    x_prix_unit = 155 * mm    # Colonne PRIX UNITAIRE (droite, déplacé vers gauche)
    x_montant = 180 * mm      # Colonne MONTANT (droite, déplacé vers gauche)
    
    # Zone Remise (ligne 105mm depuis haut = 192mm depuis bas)
    y_remise = 192 * mm
    
    # Zone MONTANT EN LETTRES (ligne 120mm depuis haut = 177mm depuis bas)
    y_montant_lettres_l1 = 177 * mm
    y_montant_lettres_l2 = 172 * mm  # Si multi-lignes
    x_montant_lettres = 16 * mm
    
    # Zone TOTAUX (cadre droite, lignes 118-130mm depuis haut)
    y_total_ht = 179 * mm   # 297 - 118 = 179
    y_total_tva = 173 * mm  # 297 - 124 = 173
    y_total_ttc = 167 * mm  # 297 - 130 = 167
    x_totaux_label = 148 * mm
    x_totaux_valeur = 180 * mm  # Aligné avec colonne MONTANT du tableau (POSITION FLÈCHE JAUNE)
    
    # Zone TRANSPORT (lignes 142-148mm depuis haut)
    y_chauffeur = 160 * mm     # 297 - 142 = 155 (AJUSTÉ VERS HAUT)
    y_matricule = 154 * mm     # 297 - 148 = 149 (AJUSTÉ VERS HAUT)
    x_chauffeur = 16 * mm      # Après label "Chauffeur :" (POSITION FLÈCHE ROUGE - Aligné avec montant en lettres)
    x_matricule = 16 * mm      # Après label "Matricule :" (POSITION FLÈCHE ROUGE)
    
    # Zone SIGNATURE (ligne 163mm depuis haut = 134mm depuis bas)
    y_signature = 139 * mm     # AJUSTÉ VERS HAUT
    x_signature = 16 * mm
    
    # Positions X spécifiques
    x_paiement = 155 * mm      # Valeurs cadre paiement (droite)
    x_client_cat_code = 170 * mm  # Catégorie + Code (alignés droite)
    
    # ========================================
    # ZONE EN-TÊTE
    # ========================================
    
    # N° FACTURE (Rouge, Gras)
    c.setFont(font_bold, size_title)
    c.setFillColor(red)
    c.drawString(150 * mm, y_facture_num, str(invoice_data.get('numero', '')))
    
    # ========================================
    # ZONE GICA (Informations Fixes Entreprise - GRIS)
    # ========================================
    
    c.setFont(font_bold, size_small)  # Changé en gras
    c.setFillColor(black)
    
    # Ligne 1 : NIF GICA + Copie B + Art + RC
    info_gica_1 = "Nif 000029096275512    Art 135135190093    Rc 00890002675    02316"
    c.drawString(margin_left, y_gica_ligne1, info_gica_1)
    
    # Ligne 2 : Copie B + autres infos fixes
    info_gica_2 = "Copie B n° : 001 00694 03508  000 315 61 81036 OUED SMAR"
    c.drawString(margin_left, y_gica_ligne2, info_gica_2)
    
    # Date avec jour de la semaine
    c.setFont(font_normal, size_normal)
    c.setFillColor(black)
    
    try:
        dt = datetime.strptime(invoice_data['date_facture'], '%Y-%m-%d')
        try:
            import locale
            try:
                locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            except:
                try:
                    locale.setlocale(locale.LC_TIME, 'French_France.1252')
                except:
                    locale.setlocale(locale.LC_TIME, 'fra_fra')
        except:
            pass
        
        try:
            date_with_day = dt.strftime('%A %d %B %Y')
        except:
            mois_fr = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                      'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
            jours_fr = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
            jour_semaine = jours_fr[dt.weekday()]
            date_with_day = f"{jour_semaine} {dt.day} {mois_fr[dt.month-1]} {dt.year}"
    except:
        date_with_day = invoice_data['date_facture']
    
    c.drawString(150 * mm, y_date, date_with_day)
    
    # ========================================
    # ZONE CLIENT GAUCHE
    # ========================================
    
    c.setFont(font_bold, size_normal)
    
    # Catégorie + Code Client (ex: SPA  0216EPN4) - Au dessus du mode de paiement
    categorie = invoice_data.get('client_categorie', invoice_data.get('categorie', ''))
    code_client = invoice_data.get('code_client', '')
    if categorie or code_client:
        client_cat_code = f"{categorie}  {code_client}".strip()
        c.drawString(x_paiement, y_paiement_reglement + 10*mm, client_cat_code)  # Juste au-dessus du paiement
    
    # Raison Sociale
    c.drawString(margin_left, y_client_nom, invoice_data.get('raison_sociale', ''))
    
    # Adresse
    c.drawString(margin_left, y_client_adresse, invoice_data.get('adresse', ''))
    
    # RC Client (nouvelle ligne)
    rc = invoice_data.get('rc', '')
    if rc:
        c.drawString(margin_left, y_client_rc, f"Rc {rc}".strip())
    
    # NIF Client + Article Imposition + RIN
    nif = invoice_data.get('nif', '')
    article_imp = invoice_data.get('article_imposition', '')
    if nif or article_imp:
        info_client = f"NIF {nif}   Art  TIN {article_imp}".strip()
        c.drawString(margin_left, y_client_nif, info_client)
    
    # ========================================
    # ZONE CONVENTION (Déplacée vers CADRE BLEU)
    # ========================================
    
    # Convention déplacée vers cadre bleu (voir ligne ~2398)
    
    # ========================================
    # ZONE PAIEMENT (Droite)
    # ========================================
    
    c.setFont(font_bold, size_normal)
    c.drawString(x_paiement, y_paiement_reglement, "Virement")
    c.drawString(x_paiement, y_paiement_banque, "BNA OUED SMAR")
    c.drawString(x_paiement, y_paiement_compte, "00 0018443")
    c.drawString(x_paiement, y_paiement_compte - 5*mm, "CMP/ prest n° 0")
    
    # ========================================
    # ZONE PRODUITS
    # ========================================
    
    c.setFont(font_bold, size_normal)
    y_current = y_produit_start
    
    for ligne in invoice_data['lignes']:
        # Code produit (colonne A) - Tronqué si trop long
        product_code = ligne.get('code_produit', '')
        product_nom = ligne.get('product_nom', '')
        
        if product_code:
            code_display = product_code[:10]  # Max 10 caractères
        else:
            code_display = product_nom[:10]
        
        c.drawString(x_code, y_current, code_display)
        
        # Description produit (colonne D) - Peut être long
        # Tronquer si nécessaire ou utiliser une police plus petite
        description = product_nom
        if len(description) > 40:
            # Couper intelligemment
            description = description[:37] + "..."
        c.drawString(x_description, y_current, description)
        
        # Unité
        c.drawString(x_unite, y_current, ligne.get('unite', ''))
        
        # Quantité (centré)
        qte = ligne.get('quantite', 0)
        qte_str = f"{qte:,.0f}".replace(",", " ")
        c.drawCentredString(x_quantite + 10*mm, y_current, qte_str)
        
        # Prix Unitaire (aligné à droite)
        pu = ligne.get('prix_unitaire', 0.0)
        pu_str = f"{pu:,.2f}".replace(",", " ").replace(".", ",")
        c.drawRightString(x_prix_unit + 25*mm, y_current, pu_str)
        
        # Montant (aligné à droite)
        montant = ligne.get('montant', 0.0)
        montant_str = f"{montant:,.2f}".replace(",", " ").replace(".", ",")
        c.drawRightString(x_montant + 25*mm, y_current, montant_str)
        
        y_current -= line_height
    
    # ========================================
    # ZONE REMISE (si applicable)
    # ========================================
    
    has_remise = invoice_data.get('montant_remise', 0) > 0 or any(l.get('taux_remise', 0) > 0 for l in invoice_data['lignes'])
    
    if has_remise:
        y_current -= line_height
        c.setFont(font_bold, size_normal)  # Assurer que c'est en gras
        c.drawString(x_description, y_current, "Remise")
        
        total_remise = sum(l.get('montant', 0) * l.get('taux_remise', 0) / 100 for l in invoice_data['lignes'])
        
        c.drawCentredString(x_quantite + 10*mm, y_current, "0%")
        c.drawRightString(x_montant + 25*mm, y_current, f"{total_remise:,.2f}".replace(",", " ").replace(".", ","))
        
        y_current -= line_height
    
    # ========================================
    # ZONE MONTANT EN LETTRES (CADRE BLEU)
    # ========================================
    
    # Montant en lettres (CADRE BLEU - Ligne 1)
    from utils import nombre_en_lettres
    montant_lettres = nombre_en_lettres(invoice_data['montant_ttc'])
    
    c.setFont(font_bold, size_small)
    c.drawString(x_chauffeur, y_chauffeur, montant_lettres)
    
    # Convention N° (CADRE BLEU - Ligne 2) - EN BLEU
    convention = invoice_data.get('convention', invoice_data.get('contrat_code', ''))
    c.setFillColor(HexColor('#0000FF'))  # Couleur bleue
    c.setFont(font_bold, size_normal)
    convention_text = f"Convention N° : {convention}" if convention else "Convention N° :"
    c.drawString(x_chauffeur, y_chauffeur - 6*mm, convention_text)
    c.setFillColor(black)  # Revenir au noir
    
    # ========================================
    # ZONE TRANSPORT (MOVED TO TOP)
    # ========================================
    
    c.setFont(font_bold, size_normal)
    
    # Chauffeur (décalé pour éviter chevauchement avec Convention)
    chauffeur = invoice_data.get('chauffeur', '')
    if chauffeur:
        c.drawString(x_chauffeur, y_matricule - 6*mm, f"Chauffeur : {chauffeur}")
    
    # Matricule (combiné tracteur + remorque)
    matricule_tracteur = invoice_data.get('matricule_tracteur', '')
    matricule_remorque = invoice_data.get('matricule_remorque', '')
    
    if matricule_tracteur or matricule_remorque:
        if matricule_tracteur and matricule_remorque:
            matricule_complet = f"{matricule_tracteur}/{matricule_remorque}"
        else:
            matricule_complet = matricule_tracteur or matricule_remorque
        
        c.drawString(x_matricule, y_matricule - 11*mm, f"Matricule : {matricule_complet}")
    
    # ========================================
    # ZONE SIGNATURE (MOVED BEFORE AMOUNT)
    # ========================================
    
    c.drawString(x_chauffeur, y_signature, "L'AGENT COMMERCIAL")
    
    # ========================================
    # ZONE TOTAUX (SANS MONTANT EN LETTRES)
    # ========================================
    
    y_totaux_base = y_current - 10*mm
    
    # Total HT (DROITE)
    c.setFont(font_bold, size_normal)
    c.drawRightString(x_totaux_valeur + 25*mm, y_totaux_base, 
                      f"{invoice_data['montant_ht']:,.2f}".replace(",", " ").replace(".", ","))
    
    # TVA 19%
    y_tva = y_totaux_base - 5*mm
    c.drawRightString(x_totaux_valeur + 25*mm, y_tva,
                      f"{invoice_data['montant_tva']:,.2f}".replace(",", " ").replace(".", ","))
    
    # Total TTC
    y_ttc = y_tva - 5*mm
    c.drawRightString(x_totaux_valeur + 25*mm, y_ttc,
                      f"{invoice_data['montant_ttc']:,.2f}".replace(",", " ").replace(".", ","))

    
    # ========================================
    # SAUVEGARDE
    # ========================================
    
    c.save()
    return output_path


def generate_calibration_pdf(output_path=None):
    """
    Générer PDF de Calibration pour Tests d'Alignement
    Grille de repères + marqueurs de position pour tous les champs
    À imprimer sur papier vierge et superposer avec préimprimé
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.colors import black, red, blue, green, gray
    
    if not output_path:
        directory = "Exports_PDF"
        if not os.path.exists(directory):
            os.makedirs(directory)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(directory, f"Calibration_Grid_{timestamp}.pdf")
    
    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4  # 210mm x 297mm
    
    # Grille de calibration 10mm x 10mm
    c.setStrokeColor(gray)
    c.setLineWidth(0.2)
    
    # Lignes verticales
    for x_pos in range(0, int(210/10) + 1):
        x = x_pos * 10 * mm
        if x_pos % 5 == 0:
            c.setLineWidth(0.5)
            c.setStrokeColor(black)
        else:
            c.setLineWidth(0.2)
            c.setStrokeColor(gray)
        c.line(x, 0, x, height)
        if x_pos % 2 == 0:
            c.setFont("Helvetica", 6)
            c.setFillColor(black)
            c.drawString(x + 1*mm, height - 5*mm, f"{x_pos*10}")
    
    # Lignes horizontales
    for y_pos in range(0, int(297/10) + 1):
        y = y_pos * 10 * mm
        if y_pos % 5 == 0:
            c.setLineWidth(0.5)
            c.setStrokeColor(black)
        else:
            c.setLineWidth(0.2)
            c.setStrokeColor(gray)
        c.line(0, y, width, y)
        if y_pos % 2 == 0:
            c.setFont("Helvetica", 6)
            c.setFillColor(black)
            c.drawString(2*mm, y + 1*mm, f"{y_pos*10}")
    
    # Fonction marqueur
    def draw_marker(x, y, label, color=red, size=8):
        c.setStrokeColor(color)
        c.setLineWidth(0.5)
        c.line(x - 2*mm, y, x + 2*mm, y)
        c.line(x, y - 2*mm, x, y + 2*mm)
        c.circle(x, y, 1*mm)
        c.setFont("Helvetica-Bold", size)
        c.setFillColor(color)
        c.drawString(x + 3*mm, y - 1*mm, label)
    
    # Titre
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(black)
    c.drawCentredString(width/2, height - 5*mm, "GRILLE DE CALIBRATION - PRÉIMPRIMÉ GICA")
    
    # Marqueurs de position
    draw_marker(185*mm, 273*mm, "N° Facture", red, 7)
    draw_marker(170*mm, 267*mm, "Date", blue, 7)
    
    c.setStrokeColor(green)
    c.setLineWidth(0.5)
    c.rect(15*mm, 257*mm, 180*mm, 4*mm)
    draw_marker(15*mm, 261*mm, "GICA L1", green, 6)
    
    draw_marker(170*mm, 253*mm, "Cat+Code", blue, 6)
    
    # Cadre client
    c.setStrokeColor(blue)
    c.rect(15*mm, 217*mm, 105*mm, 24*mm)
    c.setFont("Helvetica", 6)
    c.setFillColor(blue)
    c.drawString(16*mm, 241*mm, "→ Raison Sociale")
    c.drawString(16*mm, 232*mm, "→ Adresse")
    c.drawString(16*mm, 227*mm, "→ RC")
    c.drawString(16*mm, 222*mm, "→ NIS")
    
    # Cadre paiement
    c.setStrokeColor(blue)
    c.rect(125*mm, 223*mm, 85*mm, 18*mm)
    c.setFont("Helvetica", 6)
    c.drawString(126*mm, 233*mm, "→ Règlement")
    c.drawString(126*mm, 228*mm, "→ Banque")
    
    # Table produits
    c.setStrokeColor(red)
    c.setLineWidth(0.8)
    for x_col in [16, 25, 119, 138, 163, 195]:
        c.line(x_col*mm, 192*mm, x_col*mm, 202*mm)
    
    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(red)
    c.drawString(16*mm, 203*mm, "CODE")
    c.drawString(25*mm, 203*mm, "DESIGNATION")
    c.drawString(119*mm, 203*mm, "UM")
    c.drawString(138*mm, 203*mm, "QTE")
    c.drawString(163*mm, 203*mm, "PU")
    c.drawString(195*mm, 203*mm, "MONT")
    
    # Totaux
    c.setStrokeColor(red)
    c.rect(145*mm, 167*mm, 65*mm, 18*mm)
    c.setFont("Helvetica", 7)
    c.setFillColor(red)
    c.drawString(148*mm, 179*mm, "TOTAL HT")
    c.drawString(148*mm, 173*mm, "TVA")
    c.drawString(148*mm, 167*mm, "TOTAL TTC")
    
    # Transport
    draw_marker(50*mm, 155*mm, "Chauffeur", blue, 6)
    draw_marker(50*mm, 149*mm, "Matricule", blue, 6)
    
    # Signature
    draw_marker(16*mm, 134*mm, "SIGNATURE", red, 7)
    
    # Légende
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(black)
    c.drawString(10*mm, 20*mm, "LÉGENDE :")
    c.setFont("Helvetica", 7)
    c.setFillColor(red)
    c.drawString(10*mm, 16*mm, "● ROUGE : Zones critiques (±0.5mm)")
    c.setFillColor(blue)
    c.drawString(10*mm, 12*mm, "● BLEU : Informations client")
    c.setFillColor(green)
    c.drawString(10*mm, 8*mm, "● VERT : Zones flexibles")
    
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(black)
    c.drawString(90*mm, 20*mm, "INSTRUCTIONS :")
    c.setFont("Helvetica", 7)
    c.drawString(90*mm, 16*mm, "1. Imprimer sur papier A4 vierge")
    c.drawString(90*mm, 12*mm, "2. Superposer avec préimprimé GICA")
    c.drawString(90*mm, 8*mm, "3. Vérifier alignement marqueurs")
    c.drawString(90*mm, 4*mm, "4. Noter décalages (offsets)")
    
    c.save()
    return output_path
