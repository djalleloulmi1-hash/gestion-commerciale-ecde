"""
Utilities Module - Commercial Management System
Handles PDF generation, Excel exports, currency conversion, and backups
"""

import os
import sys
import shutil
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        # Determine width based on page size
        width, height = self._pagesize
        self.setFont("Helvetica", 9)
        self.drawRightString(width - 20*mm, 10*mm,
            "Page %d / %d" % (self._pageNumber, page_count))

def get_image(path, width=1*cm):
    pass # Placeholder if needed, but we use Image flowable mostly.

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ==================== CURRENCY CONVERSION ====================

def nombre_en_lettres(nombre: float) -> str:
    "Convert number to French words for Algerian Dinars and Centimes."
    
    unites = ["", "Un", "Deux", "Trois", "Quatre", "Cinq", "Six", "Sept", "Huit", "Neuf"]
    dizaines = ["", "Dix", "Vingt", "Trente", "Quarante", "Cinquante", 
                "Soixante", "Soixante-Dix", "Quatre-Vingt", "Quatre-Vingt-Dix"]
    speciales = ["Dix", "Onze", "Douze", "Treize", "Quatorze", "Quinze", 
                 "Seize", "Dix-Sept", "Dix-Huit", "Dix-Neuf"]
    
    def convert_centaines(n: int) -> str:
        """Convert number 0-999 to words"""
        if n == 0:
            return ""
        elif n < 10:
            return unites[n]
        elif n < 20:
            return speciales[n - 10]
        elif n < 100:
            d, u = divmod(n, 10)
            if d == 7 or d == 9:
                return dizaines[d - 1] + ("-" + speciales[u] if u > 0 else "-Dix")
            elif d == 8:
                return dizaines[d] + ("s" if u == 0 else "-" + unites[u])
            else:
                return dizaines[d] + ("-" + unites[u] if u > 0 else "")
        else:
            c, reste = divmod(n, 100)
            centaine = "Cent" if c == 1 else unites[c] + " Cent"
            if reste == 0 and c > 1:
                centaine += "s"
            return centaine + (" " + convert_centaines(reste) if reste > 0 else "")
    
    def convert_milliers(n: int) -> str:
        """Convert number with thousands"""
        if n < 1000:
            return convert_centaines(n)
        milliers, reste = divmod(n, 1000)
        if milliers == 1:
            mil = "Mille"
        else:
            mil = convert_centaines(milliers) + " Mille"
        return mil + (" " + convert_centaines(reste) if reste > 0 else "")
    
    def convert_millions(n: int) -> str:
        """Convert number with millions"""
        if n < 1000000:
            return convert_milliers(n)
        millions, reste = divmod(n, 1000000)
        if millions == 1:
            mil = "Un Million"
        else:
            mil = convert_milliers(millions) + (" Millions" if millions > 1 else " Million")
        return mil + (" " + convert_milliers(reste) if reste > 0 else "")
    
    # Handle negative numbers
    prefix = ""
    if nombre < 0:
        nombre = abs(nombre)
        prefix = "Moins "
        
    # Split into dinars and centimes
    dinars = int(nombre)
    centimes = int(round((nombre - dinars) * 100))
    
    if dinars == 0:
        result_dinars = "Zéro Dinar Algérien"
    else:
        result_dinars = convert_millions(dinars) + (" Dinars Algériens" if dinars > 1 else " Dinar Algérien")
    
    if centimes > 0:
        result_centimes = convert_centaines(centimes) + (" Centimes" if centimes > 1 else " Centime")
        return prefix + result_dinars + " et " + result_centimes
    else:
        return prefix + result_dinars


# ==================== PDF GENERATION ====================

def resource_path(relative_path: str) -> str:
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def check_logo_exists() -> str:
    "Check if logo exists, prioritizing logo_entete.png, then logo.png"
    # List of possible logo filenames
    logo_names = ["logo_entete.png", "logo.png", "logo_gica.png"]
    
    for name in logo_names:
        # Check normal path (dev) AND resource path (exe)
        # In dev, resource_path returns the absolute path too, so it covers both.
        path = resource_path(name)
        if os.path.exists(path):
            return path
            
        # Fallback to checking pure relative path just in case
        if os.path.exists(name):
             return name
             
    return None


def format_number(value: float, decimals: int = 2) -> str:
    """Format a number with thousand separators (space) and fixed decimals."""
    try:
        if value is None: value = 0.0
        s = f"{float(value):,.{decimals}f}"
        return s.replace(",", " ")
    except:
        return f"{0:.{decimals}f}"

def format_currency(value: float) -> str:
    """Format currency (2 decimals, space separator)"""
    return format_number(value, 2)

def parse_currency(value_str: Any) -> float:
    """Parse string with spaces to float. Handles '1 234,50' -> 1234.50"""
    if not value_str: return 0.0
    if isinstance(value_str, (int, float)): return float(value_str)
    
    try:
        # Remove thousands separator (space)
        clean = str(value_str).replace(" ", "")
        # Replace decimal comma with dot
        clean = clean.replace(",", ".")
        return float(clean)
    except:
        return 0.0

def format_quantity(value: float, unit: str) -> str:
    """Format quantity based on unit. 
    'Tonne' -> 3 decimals (e.g., 1 500.000)
    Others -> 2 decimals (e.g., 5.00)
    """
    decimals = 3 if str(unit).lower() == 'tonne' else 2
    return format_number(value, decimals)


def generate_invoice_pdf(facture_data: Dict[str, Any], filename: str):
    "Generate invoice or credit note PDF."
    doc = SimpleDocTemplate(filename, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    

    # Watermark Function
    def draw_watermark(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 60)
        canvas.setFillColorRGB(0.8, 0.2, 0.2, 0.3) # Red with opacity
        canvas.translate(10*cm, 15*cm)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "ANNULÉE")
        canvas.restoreState()

    if facture_data.get('statut') == 'ANNULEE' or facture_data.get('statut_facture') == 'Annulée':
         on_page = draw_watermark
    else:
         on_page = None

    # Header section with Logo
    logo_path = check_logo_exists()
    title = facture_data['type_document'].upper()
    
    if logo_path:
        img = Image(logo_path, width=2.5*cm, height=2.5*cm)
        img.hAlign = 'LEFT'
        header_data = [[img, Paragraph(title, title_style)]]
        header_table = Table(header_data, colWidths=[3*cm, 15*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
        ]))
        story.append(header_table)
    else:
        story.append(Paragraph(title, title_style))
        
    story.append(Spacer(1, 0.5*cm))
    
    # MOTIF ANNULATION (If applicable)
    if (facture_data.get('statut') == 'ANNULEE' or facture_data.get('statut_facture') == 'Annulée') and facture_data.get('motif_annulation'):
         motif_style = ParagraphStyle(
             'MotifRed', parent=styles['Normal'], textColor=colors.red, fontSize=12, alignment=TA_CENTER)
         story.append(Paragraph(f"<b>MOTIF D'ANNULATION :</b> {facture_data['motif_annulation']}", motif_style))
         story.append(Spacer(1, 0.5*cm))

    # Header info
    header_data = [
        [f"N. {facture_data['numero']}", f"Date: {facture_data['date_facture']}"],
    ]
    header_table = Table(header_data, colWidths=[9*cm, 9*cm])
    header_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Client info
    client_info = (
        f"<b>Client:</b> {facture_data['raison_sociale']}<br/>"
    )
    if facture_data.get('client_compte_bancaire'):
        client_info += f"<b>N° de compte bancaire :</b> {facture_data['client_compte_bancaire']}<br/>"
        
    client_info += (
        f"<b>Adresse:</b> {facture_data['adresse']}<br/>"
        f"<b>RC:</b> {facture_data['rc']} | <b>NIS:</b> {facture_data['nis']} | <b>NIF:</b> {facture_data['nif']}"
    )
    
    if facture_data.get('client_categorie'):
         client_info += f" | <i>(Catégorie: {facture_data['client_categorie']})</i>"
    
    if facture_data.get('contract_code'):
        client_info += f"<br/><b>Contrat N.:</b> {facture_data['contract_code']} | <b>Du:</b> {facture_data['contract_debut']} <b>Au:</b> {facture_data['contract_fin']}"
    story.append(Paragraph(client_info, styles['Normal']))
    story.append(Spacer(1, 0.8*cm))
    
    # Line items table
    table_data = [['Produit', 'Unité', 'Quantité', 'P.U. Init', 'Remise', 'Montant Net']]
    for ligne in facture_data['lignes']:
        # Handle cases where new columns might be missing (legacy data)
        prix_init = ligne.get('prix_initial', ligne['prix_unitaire'])
        remise = ligne.get('taux_remise', 0.0)
        
        table_data.append([
            ligne['product_nom'],
            ligne['unite'],
            format_quantity(ligne['quantite'], ligne['unite']),
            f"{prix_init:.2f}",
            f"{remise:.1f}%",
            f"{ligne['montant']:.2f}"
        ])
    
    # Add totals
    table_data.append(['', '', '', '', 'Total HT:', f"{facture_data['montant_ht']:.2f}"])
    table_data.append(['', '', '', '', 'TVA:', f"{facture_data['montant_tva']:.2f}"])
    table_data.append(['', '', '', '', 'Total TTC:', f"{facture_data['montant_ttc']:.2f}"])
    
    ligne_table = Table(table_data, colWidths=[5.5*cm, 1.5*cm, 2*cm, 3*cm, 2*cm, 3*cm])
    ligne_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -4), 1, colors.black),
        ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (4, -3), (4, -1), 'RIGHT'), # Align "Total Label"
        ('ALIGN', (5, -3), (5, -1), 'RIGHT'), # Align "Amount"
    ]))
    story.append(ligne_table)
    story.append(Spacer(1, 0.8*cm))
    
    # Amount in words
    montant_lettres = nombre_en_lettres(facture_data['montant_ttc'])
    story.append(Paragraph(f"<b>Arrêté la présente facture à la somme de :</b> {montant_lettres}", styles['Normal']))
    
    if on_page:
        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    else:
        doc.build(story)
    return filename


def generate_reception_pdf(reception_data: Dict[str, Any], filename: str):
    "Generate delivery receipt (BR) PDF."
    doc = SimpleDocTemplate(filename, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2e7d32'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Header Table (Logo Left, Title Center)
    logo_path = check_logo_exists()
    title_text = "BON DE RÉCEPTION"
    
    if logo_path:
        img = Image(logo_path, width=2.5*cm, height=2.5*cm)
        img.hAlign = 'LEFT'
        
        header_data = [[img, Paragraph(title_text, title_style)]]
        header_table = Table(header_data, colWidths=[3*cm, 15*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
        ]))
        story.append(header_table)
    else:
        story.append(Paragraph(title_text, title_style))
        
    story.append(Spacer(1, 0.5*cm))
    
    # Reception info
    info_data = [
        ['N. BR:', reception_data['numero']],
        ['Date:', reception_data['date_reception']],
        ['Chauffeur:', reception_data['chauffeur']],
        ['Mat. Tracteur:', reception_data['matricule']],
        ['Mat. Remorque:', reception_data.get('matricule_remorque', '-')],
        ['Transporteur:', reception_data['transporteur']],
        ['Lieu de livraison:', reception_data['lieu_livraison']],
    ]
    
    if reception_data.get('adresse_chantier'):
        info_data.append(['Adresse chantier:', reception_data['adresse_chantier']])
    
    unit = reception_data.get('unite', '') # Ensure get_all_receptions selects this or we pass it
    # Note: reception_data in generate_reception_pdf might come from a fetch that now includes 'unite'
    # But wait, generate_reception_pdf in ui.py (Step 107) does a JOIN with products!
    # "SELECT r.*, p.nom as product_nom FROM receptions r JOIN products p..."
    # I need to update THAT query in ui.py as well to fetch p.unite!
    # For now, let's assume it will be there.
    # Actually, let's use a safe fallback.
    
    # Correction: The Utils function just takes the dict. The SQL query in UI needs to be updated.
    
    info_data.extend([
        ['Produit:', reception_data['product_nom']],
        ['Quantité annoncée:', format_quantity(reception_data['quantite_annoncee'], unit)],
        ['Quantité reçue:', format_quantity(reception_data['quantite_recue'], unit)],
        ['Écart:', format_quantity(reception_data['ecart'], unit)],
    ])
    
    info_table = Table(info_data, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(info_table)
    
    doc.build(story)
    return filename


def generate_bordereau_pdf(bordereau_data: Dict[str, Any], paiements: List[Dict[str, Any]], filename: str):
    "Generate bank deposit voucher PDF."
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#c62828'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Header Table (Logo Left, Title Center)
    logo_path = check_logo_exists()
    title_text = "BORDEREAU DE VERSEMENT"
    
    if logo_path:
        img = Image(logo_path, width=2.5*cm, height=2.5*cm)
        img.hAlign = 'LEFT'
        
        # Landscape A4 is approx 29.7cm wide. 
        header_data = [[img, Paragraph(title_text, title_style)]]
        header_table = Table(header_data, colWidths=[3*cm, 23*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
        ]))
        story.append(header_table)
    else:
        story.append(Paragraph(title_text, title_style))

    story.append(Spacer(1, 0.3*cm))
    
    # Header info
    header_info = (
        f"<b>N.:</b> {bordereau_data['numero']} | "
        f"<b>Date:</b> {bordereau_data['date_bordereau']} | "
        f"<b>Banque:</b> {bordereau_data['banque']}"
    )
    story.append(Paragraph(header_info, styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Payments table
    table_data = [['N. Paiement', 'Date', 'Client', 'Mode', 'Référence', 'Montant']]
    for p in paiements:
        table_data.append([
            p['numero'],
            p['date_paiement'],
            p['client_nom'],
            p['mode_paiement'],
            p.get('reference', '-'),
            f"{p['montant']:.2f}"
        ])
    
    # Total row
    table_data.append(['', '', '', '', 'TOTAL:', f"{bordereau_data['montant_total']:.2f}"])
    
    payment_table = Table(table_data, colWidths=[4*cm, 3*cm, 6*cm, 3*cm, 4*cm, 3*cm])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -2), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffebee')),
        ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
    ]))
    story.append(payment_table)
    
    doc.build(story)
    return filename


# ==================== EXCEL EXPORTS ====================

def export_clients_to_excel(clients: List[Dict[str, Any]], filename: str):
    "Export clients list to Excel"
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Header style
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Raison Sociale', 'Seuil Crédit', 'Solde Initial (N-1)', 
               'Total Chèques', 'Total Versements', 'Total Virements', 
               'Total Paiements Global', 'Total Factures PPC (Net)', 'Solde Actuel']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Data rows
    for client in clients:
        ws.append([
            client['raison_sociale'],
            client['seuil_credit'],
            client['report_n_moins_1'],
            client.get('paiements_cheque', 0.0),
            client.get('paiements_versement', 0.0),
            client.get('paiements_virement', 0.0),
            client.get('paiements_global', 0.0),
            client.get('factures_net_ttc', 0.0),
            client.get('solde_actuel', 0.0)
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename


def export_factures_to_excel(factures: List[Dict[str, Any]], filename: str):
    "Export invoices to Excel"
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"
    
    # Header style
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Numéro', 'Type', 'Date', 'Client', 'Montant HT', 
               'Montant TVA', 'Montant TTC', 'Statut']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Data rows
    for facture in factures:
        ws.append([
            facture['numero'],
            facture['type_document'],
            facture['date_facture'],
            facture.get('client_nom', ''),
            facture['montant_ht'],
            facture['montant_tva'],
            facture['montant_ttc'],
            facture['statut']
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename


def export_stock_to_excel(products: List[Dict[str, Any]], filename: str):
    "Export stock status to Excel"
    wb = Workbook()
    ws = wb.active
    ws.title = "État des Stocks"
    
    # Header style
    header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Produit', 'Unité', 'Stock Actuel', 'Prix Actuel']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Data rows
    for product in products:
        ws.append([
            product['nom'],
            product['unite'],
            product['stock_actuel'],
            product['prix_actuel']
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename


# ==================== BACKUP ====================

def create_backup(db_path: str = "gestion_commerciale.db"):
    "Create timestamped backup of database"
    if not os.path.exists(db_path):
        return None
    
    # Create Backups directory if it doesn't exist
    backup_dir = "Backups"
    os.makedirs(backup_dir, exist_ok=True)
    
    # Generate timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"gestion_{timestamp}.db"
    backup_path = os.path.join(backup_dir, backup_filename)
    
    # Copy database file
    shutil.copy2(db_path, backup_path)
    
    return backup_path

# ==================== REPORTING PDF GENERATION ====================

def generate_creances_pdf(data: List[Dict[str, Any]], filename: str):
    "Generate PDF for Etat des Creances (A Terme invoices)"
    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4
    
    # Header
    logo_path = check_logo_exists()
    if logo_path:
        c.drawImage(logo_path, 30, height - 90, width=70, height=70, preserveAspectRatio=True, mask='auto', anchor='nw')
    
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height - 50, "ETAT DES CREANCES")
    
    c.setFont("Helvetica", 10)
    c.drawString(30, height - 100, f"Date: {datetime.now().strftime('%d/%m/%Y')}")
    
    # Table Header
    y = height - 130
    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, y, "Client")
    c.drawString(200, y, "Facture N.")
    c.drawString(300, y, "Date")
    c.drawRightString(width - 30, y, "Montant TTC")
    
    c.line(30, y - 5, width - 30, y - 5)
    y -= 25
    
    c.setFont("Helvetica", 10)
    total = 0.0
    
    for item in data:
        if y < 50:
            c.showPage()
            y = height - 50
            
        c.drawString(30, y, str(item['client'])[:35])
        c.drawString(200, y, str(item['numero']))
        c.drawString(300, y, str(item['date']))
        c.drawRightString(width - 30, y, f"{item['montant']:,.2f}")
        
        total += item['montant']
        y -= 20
        
    # Total
    c.line(30, y + 10, width - 30, y + 10)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(300, y - 10, "TOTAL CREANCES:")
    c.drawRightString(width - 30, y - 10, f"{total:,.2f} DA")
    
    c.save()

def generate_ca_pdf(data: Dict[str, Any], filename: str):
    "Generate PDF for Etat du Chiffre d Affaires"
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=4.5*cm)
    story = []
    width, height = A4
    styles = getSampleStyleSheet()

    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ETAT DU CHIFFRE D'AFFAIRES")
        
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Période: {data['start_date']} au {data['end_date']}")
        
        canvas.restoreState()

    # story.append(Spacer(1, 4*cm))

    # Summary Table
    # Layout:
    # CA BRUT | value
    # TOTAL AVOIRS | value
    # CA NET | value
    
    summary_data = [
        ["CA BRUT (Factures):", f"{data['ca_brut']:,.2f} DA"],
        ["TOTAL AVOIRS:", f"- {data['total_avoirs']:,.2f} DA"],
        ["CA NET:", f"{data['ca_net']:,.2f} DA"]
    ]
    
    summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'), # CA NET Bold
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 1*cm))
    
    # Details of Avoirs
    if data['details_avoirs']:
        story.append(Paragraph("Détail des Avoirs déduits:", styles['Heading3']))
        story.append(Spacer(1, 0.2*cm))
        
        headers = ["Avoir N.", "Ref Facture", "Date", "Montant HT"]
        table_data = [headers]
        
        for item in data['details_avoirs']:
            table_data.append([
                str(item['numero']),
                str(item['facture_ref']),
                str(item['date']),
                f"{item['montant']:,.2f}"
            ])
            
        t = Table(table_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Aucun avoir sur cette période.", styles['Normal']))

    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)


def generate_situation_pdf(data: Dict[str, Any], filename: str):
    "Generate PDF for Client Situation"
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=4.5*cm)
    story = []
    width, height = A4
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "SITUATION CLIENT")
        
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Date: {datetime.now().strftime('%d/%m/%Y')}")
        
        canvas.restoreState()
    
    # story.append(Spacer(1, 4*cm))
    
    client = data['client']
    story.append(Paragraph(f"<b>Client:</b> {client['raison_sociale']}", styles['Normal']))
    story.append(Paragraph(f"<b>Adresse:</b> {client['adresse']}", styles['Normal']))
    story.append(Spacer(1, 1*cm))
    
    # Balance Summary
    balance = data['balance']
    
    summary_data = [
        ["Résumé Financier", ""],
        ["Report N-1:", f"{balance['report']:,.2f}"],
        ["Total Factures:", f"{balance['total_factures']:,.2f}"],
        ["Total Paiements:", f"{balance['total_paiements']:,.2f}"],
        ["Total Avoirs:", f"{balance['total_avoirs']:,.2f}"],
        ["SOLDE ACTUEL:", f"{balance['solde']:,.2f} DA"]
    ]
    
    t = Table(summary_data, colWidths=[8*cm, 6*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (1, 0)), # Merge title
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey), # Header bg
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        
        # Solde Row
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('TOPPADDING', (0, -1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
    ]))
    
    story.append(t)

    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)

def generate_client_state_pdf(clients: List[Dict[str, Any]], filename: str):
    "Generate PDF for detailed client state"
    # Use landscape because there are many columns
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4), topMargin=4.5*cm)
    story = []
    
    width, height = landscape(A4)
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ÉTAT DES CLIENTS DÉTAILLÉ")
        
        # Subtitle/Date
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Date: {datetime.now().strftime('%d/%m/%Y')}")
        
        # Line
        canvas.setLineWidth(1)
        # canvas.line(1*cm, height - 3.8*cm, width - 1*cm, height - 3.8*cm)
        
        canvas.restoreState()
    
    # Spacer for header
    # story.append(Spacer(1, 4*cm))
    
    # Table Data
    # Columns: Raison Soc, Adresse, RC, NIS, NIF, Art Imp
    headers = ['Raison Sociale', 'Adresse', 'N. RC', 'N. NIS', 'N. NIF', 'N. Art. Imp']
    table_data = [headers]
    
    for client in clients:
        table_data.append([
            client['raison_sociale'],
            client['adresse'],
            client['rc'],
            client['nis'],
            client['nif'],
            client['article_imposition']
        ])
        
    # Create Table
    # Adjust column widths for landscape A4 (approx 28cm printable width)
    col_widths = [6*cm, 7*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Align left for text data
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('rowBackground', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]), # Alternating colors
    ]))
    
    story.append(t)
    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return filename


def generate_invoice_state_pdf(invoice_lines: List[Dict[str, Any]], date_range: Dict[str, str], filename: str):
    "Generate PDF for detailed invoice state by date range"
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=4.5*cm)
    story = []
    
    width, height = A4
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ÉTAT DES FACTURES")
        
        # Subtitle/Date
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Période: {date_range['start']} au {date_range['end']}")
        
        canvas.restoreState()
    
    # Spacer for header
    # story.append(Spacer(1, 4*cm))
    
    # Table Data
    # Columns: N. Facture, Date, Produit, Montant HT
    headers = ['N. Facture', 'Date', 'Produit', 'Montant HT']
    table_data = [headers]
    
    total_ht = 0.0
    
    # Track which rows are Avoirs to style them
    avoir_rows = []
    
    for idx, line in enumerate(invoice_lines):
        # Determine if Avoir or Cancelled.
        is_avoir = (line.get('type_document') == 'Avoir')
        is_cancelled = (line.get('statut') == 'Annulée')
        has_avoir = line.get('has_avoir', 0) # SQLite return 0 or 1
        
        is_highlighted = is_avoir or is_cancelled or has_avoir
        
        table_data.append([
            line['numero'],
            line['date_facture'],
            line['product_nom'],
            f"{line['montant_ht']:,.2f}"
        ])
        
        # Only add to total if NOT cancelled
        if not is_cancelled:
            total_ht += line['montant_ht']
        
        if is_highlighted:
            avoir_rows.append(idx + 1) # +1 for header
        
    # Total Row
    table_data.append(['', '', 'TOTAL HT:', f"{total_ht:,.2f}"])
        
    # Create Table
    col_widths = [4*cm, 3*cm, 8*cm, 4*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Base Style
    base_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'), # Amount right align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'), # Changed to Bold
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Total Row Style
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
    ]
    
    # Add Red Color for Avoirs
    for row_idx in avoir_rows:
        base_style.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.darkred))
    
    t.setStyle(TableStyle(base_style))
    
    story.append(t)
    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return filename


def generate_etat_104_pdf(sales_data: List[Dict[str, Any]], date_range: Dict[str, str], filename: str):
    "Generate PDF for Etat 104 (Sales by Client)"
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4), topMargin=4.5*cm)
    story = []
    
    width, height = landscape(A4)
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ÉTAT N. 104 - CHIFFRE D'AFFAIRES PAR CLIENT")
        
        # Subtitle/Date
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Période: {date_range['start']} au {date_range['end']}")
        
        canvas.restoreState()
    
    # Spacer for header
    # story.append(Spacer(1, 4*cm))
    
    # Table Data
    # Columns: Raison Sociale, RC, NIF, NIS, Art Imp, CA HT
    headers = ['Raison Sociale', 'N. RC', 'N. NIF', 'N. NIS', 'Art. Imposition', 'CA HT']
    table_data = [headers]
    
    total_ca = 0.0
    
    for row in sales_data:
        table_data.append([
            row['raison_sociale'],
            row['rc'],
            row['nif'],
            row['nis'],
            row['article_imposition'],
            f"{row['chiffre_affaire_ht']:,.2f}"
        ])
        total_ca += row['chiffre_affaire_ht']
        
    # Total Row
    table_data.append(['', '', '', '', 'TOTAL CA HT:', f"{total_ca:,.2f}"])
        
    # Create Table
    col_widths = [6*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 4*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'), # Amount right align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Total Row Style
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
    ]))
    
    story.append(t)
    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return filename


def generate_payments_state_pdf(payments: List[Dict[str, Any]], date_range: Dict[str, str], filename: str):
    "Generate PDF for detailed payments state by date range"
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=4.5*cm)
    story = []
    
    width, height = A4
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ÉTAT DES PAIEMENTS")
        
        # Subtitle/Date
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Période: {date_range['start']} au {date_range['end']}")
        
        canvas.restoreState()
    
    # Spacer for header
    # story.append(Spacer(1, 4*cm))
    
    # Table Data
    # Columns: N. Paiement, Date, Client, Mode, Réf, Montant
    headers = ['N. Paiement', 'Date', 'Client', 'Mode', 'Réf', 'Montant']
    table_data = [headers]
    
    total = 0.0
    
    for p in payments:
        table_data.append([
            p['numero'],
            p['date_paiement'],
            p['client_nom'],
            p['mode_paiement'],
            p.get('reference', '-'),
            f"{p['montant']:.2f}"
        ])
        total += p['montant']
        
    # Total Row
    table_data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}"])
        
    # Create Table
    col_widths = [3.5*cm, 2.5*cm, 5*cm, 2.5*cm, 2.5*cm, 3*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6')),
    ]))
    
    story.append(t)
    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return filename



def get_conditional_styles(data_matrix, start_row=0, start_col=0):
    """
    Generate ReportLab TableStyle commands for conditional formatting.
    Positives (>0) -> Green
    Negatives (<0) -> Orange (#ff9800)
    Zeros -> Blue
    """
    styles = []
    orange_color = colors.HexColor('#ff9800')
    green_color = colors.green
    blue_color = colors.blue
    valid_chars = "0123456789.-"
    
    for r_idx, row in enumerate(data_matrix):
        for c_idx, cell_value in enumerate(row):
            cell_str = str(cell_value).replace(" DA", "").replace("%", "")
            
            # If contains letters, it is text (Description, Client Name, etc.) -> Skip
            if any(c.isalpha() for c in cell_str):
                continue

            # Val cleanup: Keep only numbers, dots and negative sign
            val_str = "".join([c for c in cell_str if c in valid_chars])
            
            try:
                # Check for empty string or non-numeric first
                if not val_str or val_str == "." or val_str == "-":
                    continue

                val = float(val_str)
                actual_row = r_idx + start_row
                actual_col = c_idx + start_col
                
                # Check for zero (handling tiny approximations)
                if abs(val) < 0.001:
                     styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), blue_color))
                elif val > 0:
                    styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), green_color))
                else:
                    styles.append(('TEXTCOLOR', (actual_col, actual_row), (actual_col, actual_row), orange_color))

            except (ValueError, TypeError):
                # Text or other non-numeric content -> Default Black
                pass
                
    return styles

def generate_daily_sales_pdf(data: Dict[str, Any], filename: str):
    "Generate Daily Sales Report PDF (Etat de vente journalier)"
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4), topMargin=4.5*cm)
    story = []
    
    width, height = landscape(A4)
    
    styles = getSampleStyleSheet()
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo (Top Left)
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            # Draw Image: x=1cm, y=height-3cm, w=3.5cm, h=2cm
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass

        # 2. Text (Center/Left aligned next to logo)
        # We use a Paragraph to handle formatting (bold, size) easily, but drawing Paragraph on canvas requires 'wrapOn' and 'drawOn'
        
        # Center X position for title
        center_x = width / 2.0
        
        # Manually drawing text for header using standard fonts to keep it simple and consistent
        # Title 1
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        # Title 2
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        # Main Title
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ETAT DES VENTES QUOTIDIENNES")
        
        # Subtitle
        canvas.setFont("Helvetica-Oblique", 10)
        canvas.drawCentredString(center_x, height - 3.0*cm, "Dépôt Oued Smar")
        
        # Date (Top Right or Below Title)
        canvas.setFont("Helvetica", 11)
        # canvas.drawRightString(width - 1*cm, height - 3.5*cm, f"Journée du : {data['date']}")
        canvas.drawCentredString(center_x, height - 3.5*cm, f"Journée du : {data['date']}")
        
        # Bottom Line of Header
        canvas.setLineWidth(1)
        canvas.line(1*cm, height - 3.8*cm, width - 1*cm, height - 3.8*cm)
        
        canvas.restoreState()

    # --- Build Story (Body) ---
    
    # Spacer for Header (since header is drawn on canvas, we need to push body down)
    # Header takes approx 4cm
    # story.append(Spacer(1, 2.5*cm))
        
    # 2. Detailed Table (Top)
    # COLS: Code, Clients, Produit (Code), Facture (N., Date), Qte, HT, MR(?), TVA, TTC
    
    headers = [
        'Code', 'Clients', 
        'Produit\nCode', 
        'Facture\nN.', 'Facture\nDate',
        'Qte', 'HT', 'MR', 'TVA', 'TTC'
    ]
    
    # Prepare data rows
    table_data = [headers]
    
    for row in data['details']:
        # Format date dd-mmm-yy
        try:
            d_obj = datetime.strptime(row['date'], '%Y-%m-%d')
            date_fmt = d_obj.strftime('%d-%b-%y')
        except:
            date_fmt = row['date']
            
        table_data.append([
            row['code_client'] or '',
            Paragraph(row['client'][:25], styles['Normal']), # Truncate/Wrap client name
            row['code_produit'] or row['produit'][:10], 
            row['facture_num'],
            date_fmt,
            f"{row['qte']:,.3f}",
            f"{row['ht']:,.2f}",
            "0,00", # MR placeholder
            f"{row['tva']:,.2f}",
            f"{row['ttc']:,.2f}"
        ])
    
    # Add Total Row for Details
    totals = data['totals']
    table_data.append([
        '', '', '', '', 'TOTAL',
        f"{totals['day_qty']:,.3f}",
        f"{totals['day_ht']:,.2f}",
        "0,00",
        f"{totals['day_tva']:,.2f}",
        f"{totals['day_ttc']:,.2f}"
    ])
    
    # Table Style
    # Revised widths to prevent overlap
    # Total available width approx 27-28cm
    det_col_widths = [2.2*cm, 5.5*cm, 2.8*cm, 3.5*cm, 2.0*cm, 2.0*cm, 2.5*cm, 1.5*cm, 2.5*cm, 3.0*cm]
    det_table = Table(table_data, colWidths=det_col_widths, repeatRows=1)
    
    det_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'), # Global Bold
        ('FONTSIZE', (0, 0), (-1, 0), 7), # Reduced header size
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'), # Numbers right aligned
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 7), # Reduced body size
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        # ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), # Covered by global
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ] + get_conditional_styles(table_data[1:], start_row=1, start_col=0)))
    
    story.append(det_table)
    story.append(Spacer(1, 1*cm))
    
    # 3. Summary Table (Bottom)
    
    product_stats = data['product_stats']
    
    # New Layout: Rows = Products, Cols = Designation, Qte Journée, Qte Cumulée
    
    # Headers
    summary_headers = ['Désignation', 'Qte Journée', 'Qte Cumulée']
    summary_data = [summary_headers]
    
    for p in product_stats:
        row = [
            p['nom'],
            f"{p['daily_qty']:,.3f}" if p['daily_qty'] != 0 else "-",
            f"{p['cumul_qty']:,.3f}" if p['cumul_qty'] != 0 else "-"
        ]
        summary_data.append(row)
        
    # Column widths
    # Total width approx 28cm
    sum_col_widths = [12*cm, 5*cm, 5*cm]
    
    sum_table = Table(summary_data, colWidths=sum_col_widths)
    # Don't repeat rows for summary table if it splits (though it shouldn't usually)
    
    sum_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'), # Global Bold
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'), # Align product names left
        ('LEFTPADDING', (0, 1), (0, -1), 5),
        
        # Header Style
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        
        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.aliceblue]),
    ] + get_conditional_styles(summary_data[1:], start_row=1, start_col=0)))
    
    story.append(sum_table)
    story.append(Spacer(1, 1*cm))
    
    # 4. Footer Totals
    # TOTAL ANNEE | QTE | [Year Qty??] | HT | [Year Net HT]
    
    footer_data = [
        ['TOTAL ANNEE', 'HT', f"{totals['year_net_ht']:,.2f} DA"]
    ]
    footer_table = Table(footer_data, colWidths=[4*cm, 2*cm, 6*cm])
    footer_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
    ]))
    
    story.append(footer_table)
    
    # Signature
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph("LE CHEF SERVICE COMMERCIAL", ParagraphStyle('Sig', parent=styles['Normal'], alignment=TA_RIGHT, fontName='Helvetica-Bold')))
    
    # Build with NumberedCanvas
    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)
    return filename


def generate_sales_by_category_pdf(data: Dict[str, Any], start_date: str, end_date: str, filename: str):
    """
    Generate PDF for periodic turnover by category (Ciment vs Others)
    """
    doc = SimpleDocTemplate(filename, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=4.5*cm, bottomMargin=2*cm)
    
    width, height = A4
    elements = []
    
    # --- Header Callback ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # 1. Logo
        logo_path = "logo_gica.png" if os.path.exists("logo_gica.png") else check_logo_exists()
        if logo_path:
            try:
                canvas.drawImage(logo_path, 1*cm, height - 3*cm, width=3.5*cm, height=2.0*cm, preserveAspectRatio=True, mask='auto')
            except:
                pass
        
        # 2. Text
        center_x = width / 2.0
        
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(center_x, height - 1.2*cm, "GROUPE INDUSTRIEL DES CIMENTS D'ALGERIE")
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, height - 1.7*cm, "ENTREPRISE DES CIMENTS ET DERIVES D'ECH-CHELIFF")
        
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(center_x, height - 2.5*cm, "ÉTAT DU CHIFFRE D'AFFAIRES PÉRIODIQUE")
        
        # Subtitle/Date
        canvas.setFont("Helvetica", 11)
        canvas.drawCentredString(center_x, height - 3.2*cm, f"Période: {start_date} au {end_date}")
        
        canvas.restoreState()
    
    # Spacer for header
    # elements.append(Spacer(1, 4*cm))
    
    styles = getSampleStyleSheet()
    
    # Define Table Style
    # Define Table Style (Global Bold)
    base_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.3, 0.4)), # Dark blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'), # Global Bold
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white), # White rows
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'), # Align Names Left
        ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'), # Align Numbers Right
    ]

    subtotal_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('Font', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    grand_total_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    grand_total = 0.0
    
    # Priority Categories
    categories = ["Ciment"] + [k for k in data.keys() if k != "Ciment" and data[k]]
    
    for cat in categories:
        if cat not in data or not data[cat]: continue
        
        # Category Header
        elements.append(Paragraph(f"<b>FAMILLE: {cat.upper()}</b>", 
                                 ParagraphStyle('CatHeader', parent=styles['Heading2'], fontSize=12, spaceAfter=10, textColor=colors.darkblue)))
        
        # Table Data
        # Header Row
        table_data = [["Désignation", "Unité", "Quantité", "Montant HT"]]
        
        cat_total = 0.0
        
        for item in data[cat]:
            row = [
                item['nom'],
                item['unite'],
                f"{item['qte']:,.2f}",
                f"{item['montant_ht']:,.2f}"
            ]
            table_data.append(row)
            cat_total += item['montant_ht']
            
        grand_total += cat_total
        
        # Create Table
        t = Table(table_data, colWidths=[8*cm, 3*cm, 3*cm, 4*cm])
        t.setStyle(TableStyle(base_style_cmds + get_conditional_styles(table_data[1:], start_row=1, start_col=0)))
        elements.append(t)
        
        # Subtotal
        sub_data = [[f"S-TOTAL {cat.upper()}", "", "", f"{cat_total:,.2f}"]]
        st = Table(sub_data, colWidths=[8*cm, 3*cm, 3*cm, 4*cm])
        st.setStyle(subtotal_style)
        elements.append(st)
        
        elements.append(Spacer(1, 15))
        
    # Grand Total
    elements.append(Spacer(1, 10))
    gt_data = [["TOTAL GÉNÉRAL HT", f"{grand_total:,.2f} DA"]]
    gt = Table(gt_data, colWidths=[14*cm, 4*cm])
    gt.setStyle(grand_total_style)
    elements.append(gt)
    
    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header, canvasmaker=NumberedCanvas)


def preview_and_print_pdf(filename: str):
    """Open PDF for preview (and printing via viewer)"""
    import os
    if os.name == 'nt':
        try:
            os.startfile(filename)
        except Exception as e:
            print(f"Error opening PDF: {e}")
    else:
        # Fallback for other OS (dev environment)
        import subprocess
        try:
            if os.uname().sysname == "Darwin":
                subprocess.call(["open", filename])
            else:
                subprocess.call(["xdg-open", filename])
        except:
            pass
